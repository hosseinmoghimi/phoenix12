from .models import Vehicle,ServiceMan,Maintenance,VehicleStatus,WorkShift,Driver
from .apps import APP_NAME
from .enums import *
from log.repo import LogRepo 
from django.db.models import Q
from django.shortcuts import reverse
from authentication.repo import PersonRepo
from utility.constants import FAILED,SUCCEED
from utility.log import leolog
from utility.calendar import PersianCalendar
from .models import VehicleEvent,Tavaghof,VehicleStatus
from .models import FilterService,OilService,Tavaghof,Product,AnbarProduct,Service
from .constants import EXCEL_VEHICLES_DATA_START_ROW
class VehicleStatusRepo():
    def __init__(self,request,*args, **kwargs):
        self.me=None
        self.my_accounts=[]
        self.request=request
        self.objects=VehicleStatus.objects.filter(id=0)
        profile=PersonRepo(request=request).me
        if profile is not None:
            if request.user.has_perm(APP_NAME+".view_vehicle"):
                self.objects=VehicleStatus.objects
                self.my_accounts=self.objects 
    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            objects=objects.filter(Q(name__contains=search_for) | Q(code=search_for)  )
        if "parent_id" in kwargs:
            parent_id=kwargs["parent_id"]
            objects=objects.filter(parent_id=parent_id)  
        if "vehicle_id" in kwargs and kwargs['vehicle_id']:
            vehicle_id=kwargs["vehicle_id"]
            objects=objects.filter(vehicle_id=vehicle_id)  
        return objects.all().order_by('-status_datetime')
        
    def vehicle_status(self,*args, **kwargs):
        if "vehicle_status_id" in kwargs and kwargs["vehicle_status_id"] is not None:
            return self.objects.filter(pk=kwargs['vehicle_status_id']).first()  
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        
        
    def add_vehicle_status(self,*args,**kwargs):
        result,message,vehicle_status=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_vehicle"):
            message="دسترسی غیر مجاز"
            return result,message,vehicle_status

        vehicle_status=VehicleStatus()
        if 'cooler' in kwargs:
            vehicle_status.cooler=kwargs["cooler"]

            
        if 'vehicle_id' in kwargs:
            vehicle_status.vehicle_id=kwargs["vehicle_id"]

        if 'status_datetime' in kwargs:
            vehicle_status.status_datetime=kwargs["status_datetime"]
        if 'wiring' in kwargs:
            vehicle_status.wiring=kwargs["wiring"]
        if 'heater' in kwargs:
            vehicle_status.heater=kwargs["heater"]

        if 'cabin' in kwargs:
            vehicle_status.cabin=kwargs["cabin"]

        if 'ziroband' in kwargs:
            vehicle_status.ziroband=kwargs["ziroband"]

                
        if 'hydrolic' in kwargs:
            vehicle_status.hydrolic=kwargs["hydrolic"]
        if 'motor' in kwargs:
            vehicle_status.motor=kwargs["motor"]
        if 'hour' in kwargs:
            vehicle_status.hour=kwargs["hour"]
        
                
        if 'description' in kwargs:
            vehicle_status.description=kwargs["description"]
        if 'location' in kwargs:
            vehicle_status.location=kwargs["location"]
        if 'pakat' in kwargs:
            vehicle_status.pakat=kwargs["pakat"]
        if 'compress' in kwargs:
            vehicle_status.compress=kwargs["compress"]

        vehicle_status.kilometer=0
        if 'kilometer' in kwargs and kwargs["kilometer"]:
                    kilometer=0
                    kilometer=kwargs["kilometer"]
                    try:
                        kilometer=int(kilometer)
                    except:
                        pass
                    vehicle_status.kilometer=kilometer

        vehicle_status.hour=0
        if 'hour' in kwargs and kwargs["hour"]:
                    hour=0
                    hour=kwargs["hour"]
                    try:
                        hour=float(hour)
                    except:
                        pass
                    vehicle_status.hour=hour
            
        (result,message,vehicle_status)=vehicle_status.save()
        return result,message,vehicle_status

    def last_statuses(self,*args, **kwargs):
        vehicles=Vehicle.objects.all()
        ids=[]
        for vehicle in vehicles:
            stat=VehicleStatus.objects.filter(vehicle_id=vehicle.id).order_by('-status_datetime').first()
            if stat is not None:
                ids.append(stat.id)
        
        return VehicleStatus.objects.filter(id__in=ids)

 
    def import_vehicle_statuses_from_excel(self,*args,**kwargs):
        result,message,vehicle_statuses=FAILED,"",[]
        excel_file=kwargs['excel_file']
         
        import openpyxl 

        wb = openpyxl.load_workbook(excel_file)
        try:
            ws = wb['vehicle_statuses']
            

        except:
            message='فایل شما  وضعیت ماشین آلات ندارد.'
            return result,message,None
        count=kwargs['count']
        try:
            count=int(ws.cell(row=1, column=2).value)
        except:
            
            message='فایل برگه وضعیت ماشین آلات ، تعداد ندارد.'
            return result,message,None 

        vehicle_statuses_to_import=[]

        START_ROW=EXCEL_VEHICLES_DATA_START_ROW
        for i in range(START_ROW,count+START_ROW):
            vehicle_status={}
            
            i=str(i) 
            # product['id']=ws['A'+str(i)].value
            vehicle_code=ws['B'+i].value
            from django.utils import timezone

            if vehicle_code is not None:
                id=0
                vehicle=Vehicle.objects.filter(vehicle_code=vehicle_code).first()
                if vehicle is not None:
                    vehicle_status['vehicle_id']=vehicle.id

                    vehicle_status['status_datetime']=(ws['D'+i].value)
                    vehicle_status['status_datetime']=timezone.now()

                    vehicle_status['location']=(ws['E'+i].value)

                    kilometer=(ws['F'+i].value)
                    try:
                        kilometer=int(kilometer)
                    except:
                        pass
                    vehicle_status['kilometer']=kilometer

                    hour=(ws['G'+i].value)
                    try:
                        hour=float(hour)
                    except:
                        pass
                    vehicle_status['hour']=hour

                    vehicle_status['motor']=(ws['H'+i].value)
                    vehicle_status['ziroband']=(ws['I'+i].value)
                    vehicle_status['cabin']=(ws['J'+i].value)
                    vehicle_status['compress']=(ws['K'+i].value)
                    vehicle_status['hydrolic']=(ws['L'+i].value)
                    vehicle_status['pakat']=(ws['M'+i].value)
                    vehicle_status['cooler']=(ws['N'+i].value)
                    vehicle_status['heater']=(ws['O'+i].value)
                    vehicle_status['gear_box']=(ws['P'+i].value)
                    vehicle_status['wiring']=(ws['Q'+i].value)
                    vehicle_status['light']=(ws['R'+i].value)
                    vehicle_status['description']=(ws['S'+i].value)
                  
                     
                    # vehicle_status['thumbnail_origin']=ws['F'+str(i)].value
                    vehicle_statuses_to_import.append(vehicle_status) 
        modified=added=0

        i=0
        for vehicle_status in vehicle_statuses_to_import:
            result,message,new_vehicle_status=self.add_vehicle_status(**vehicle_status)
        
            if result==SUCCEED:
                added+=1

        result=SUCCEED
        message=f"""{added} وضعیت دستگاه اضافه شد."""
        vehicle_status=self.list()

        
        if True:
            log_data={}
            from log.repo import LogRepo
            log_data['person_id']=PersonRepo(request=self.request).me.id
            log_data['url']=reverse("transport:vehicle_statuses")
            log_data['title']="بازیابی سرویس ها"
            log_data['description']=message
            log_data['app_name']=APP_NAME
            LogRepo(request=self.request).add_log(**log_data)

        return result,message,vehicle_status


class WorkShiftRepo():
    def __init__(self,request,*args, **kwargs):
        self.me=None
        self.my_accounts=[]
        self.request=request
        self.objects=WorkShift.objects.filter(id=0)
        profile=PersonRepo(request=request).me
        if profile is not None:
            if request.user.has_perm(APP_NAME+".view_vehicle"):
                self.objects=WorkShift.objects
                self.my_accounts=self.objects 

    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            objects=objects.filter(Q(name__contains=search_for) | Q(code=search_for)  )
         
        if "shift" in kwargs and kwargs["shift"]:
            objects=objects.filter(shift=kwargs["shift"])

        if "location" in kwargs and kwargs["location"]:
            objects=objects.filter(location=kwargs["location"]) 
        if "vehicle_id" in kwargs:
            vehicle_id=kwargs["vehicle_id"]
            objects=objects.filter(vehicle_id=vehicle_id)  

        if "vehicle_code" in kwargs and kwargs['vehicle_code']:
            objects=objects.filter(vehicle__vehicle_code=kwargs["vehicle_code"])  
        if "driver_id" in kwargs and kwargs["driver_id"]:
            objects=objects.filter(driver_id=kwargs["driver_id"])  
        if "shift_date" in kwargs and kwargs['shift_date']:
            year=kwargs['shift_date'][:2]
            if year=="13" or year=="14":
                kwargs['shift_date']=PersianCalendar().to_gregorian(kwargs["shift_date"])
                kwargs['shift_date']=kwargs['shift_date'].date()
            objects=objects.filter(shift_date=kwargs["shift_date"]) 

        if "from_shift_date" in kwargs and kwargs['from_shift_date']:
            year=kwargs['from_shift_date'][:2]
            if year=="13" or year=="14":
                kwargs['from_shift_date']=PersianCalendar().to_gregorian(kwargs["from_shift_date"])
            
            objects=objects.filter(shift_date__gte=kwargs["from_shift_date"]) 

        if "to_shift_date" in kwargs and kwargs['to_shift_date']:
            year=kwargs['to_shift_date'][:2]
            if year=="13" or year=="14":
                kwargs['to_shift_date']=PersianCalendar().to_gregorian(kwargs["to_shift_date"])

            import datetime 
            delta=datetime.timedelta(hours=23,minutes=59,seconds=59)
            kwargs['to_shift_date']=kwargs['to_shift_date']+delta                        
            objects=objects.filter(shift_date__lte=kwargs["to_shift_date"]) 

            
        if "shift" in kwargs and kwargs['shift']:
            objects=objects.filter(shift=kwargs["shift"])  
        return objects.all()
        
    def work_shift(self,*args, **kwargs):
        if "work_shift_id" in kwargs and kwargs["work_shift_id"] is not None:
            return self.objects.filter(pk=kwargs['work_shift_id']).first()  
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        
        
    def add_work_shift(self,*args,**kwargs):

        result,message,work_shift=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_vehicle"):
            message="دسترسی غیر مجاز"
            return result,message,work_shift

        work_shift=WorkShift()
        if 'title' in kwargs:
            work_shift.title=kwargs["title"]
            if len(Vehicle.objects.filter(title=work_shift.title))>0:
                message='نام تکراری برای وسیله نقلیه جدید'
                return FAILED,message,None
        if 'vehicle_code' in kwargs:
            vehicle_code=kwargs["vehicle_code"]
            vehicle=Vehicle.objects.filter(vehicle_code=vehicle_code).first()
            if vehicle is not None:
                work_shift.vehicle=vehicle

        if 'driver_id' in kwargs:
            work_shift.driver_id=kwargs["driver_id"]

        if 'location' in kwargs:
            work_shift.location=kwargs["location"]

        if 'shift_date' in kwargs:
            year=kwargs['shift_date'][:2]
            if year=="13" or year=="14":
                kwargs['shift_date']=PersianCalendar().to_gregorian(kwargs["shift_date"])
            work_shift.shift_date=kwargs["shift_date"]

        if 'shift' in kwargs:
            work_shift.shift=kwargs["shift"]
        if 'start_hour' in kwargs:
            work_shift.start_hour=kwargs["start_hour"]
        if 'end_hour' in kwargs:
            work_shift.end_hour=kwargs["end_hour"]

        if 'vehicle_start_hour' in kwargs:
                    work_shift.vehicle_start_hour=kwargs["vehicle_start_hour"]

        if 'vehicle_end_hour' in kwargs:
                    work_shift.vehicle_end_hour=kwargs["vehicle_end_hour"]


        if 'bar_count' in kwargs:
            work_shift.bar_count=kwargs["bar_count"]

        if 'bar' in kwargs:
            work_shift.bar=kwargs["bar"]

        if 'description' in kwargs:
            work_shift.description=kwargs["description"]
 
 
        if 'gasoil_liter' in kwargs:
            work_shift.gasoil_liter=kwargs["gasoil_liter"]
          
         

        
        (result,message,work_shift)=work_shift.save()

        if 'filters' in kwargs and result==SUCCEED:
            filters=kwargs["filters"] 
            for filter_ in filters:
                filter_service=FilterService()
                filter_service.filter_type=filter_['filter_type']
                filter_service.filter_action=filter_['filter_action']
                filter_service.count=filter_['count']
                filter_service.cost=filter_['cost']
                filter_service.description=filter_['description']
                filter_service.work_shift=work_shift
                filter_service.save()

                
        if 'oils' in kwargs and result==SUCCEED:
            oils=kwargs["oils"] 
            for oil_ in oils:
                oil_service=OilService()
                oil_service.oil_type=oil_['oil_type']
                oil_service.oil_action=oil_['oil_action']
                oil_service.oil_liter=oil_['oil_liter']
                oil_service.cost=oil_['cost']
                oil_service.description=oil_['description']
                oil_service.work_shift=work_shift
                oil_service.save()

                
        if 'tavaghofs' in kwargs and result==SUCCEED:
            tavaghofs=kwargs["tavaghofs"] 
            for tavaghof_ in tavaghofs:
                tavaghof=Tavaghof()
                tavaghof.cause=tavaghof_['cause']
                tavaghof.duration=tavaghof_['duration']
                tavaghof.description=tavaghof_['description']
                tavaghof.vehicle_hour=tavaghof_['vehicle_hour']
                tavaghof.work_shift=work_shift
                tavaghof.save()
 
                
        if 'products' in kwargs and result==SUCCEED:
            products=kwargs["products"] 
            for product_ in products:
                product=Product()
                product.name=product_['name']
                product.quantity=product_['quantity']
                product.description=product_['description']
                product.unit_price=product_['unit_price']
                product.service_man=product_['service_man']
                product.anbar=product_['anbar']
                product.work_shift=work_shift
                product.save()
 
                

        return result,message,work_shift


class OilServiceRepo():
    def __init__(self,request,*args, **kwargs):
        self.me=None
        self.my_accounts=[]
        self.request=request
        self.objects=OilService.objects.filter(id=0)
        profile=PersonRepo(request=request).me
        if profile is not None:
            if request.user.has_perm(APP_NAME+".view_oil_service"):
                self.objects=OilService.objects
                self.my_accounts=self.objects 
    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            objects=objects.filter(Q(name__contains=search_for) | Q(code=search_for)  )
        if "vehicle_id" in kwargs:
            vehicle_id=kwargs["vehicle_id"]
            objects=objects.filter(work_shift__vehicle_id=vehicle_id)  
        if "owner_id" in kwargs:
            owner_id=kwargs["owner_id"]
            objects=objects.filter(owner_id=owner_id)  
        return objects.all()
        
    def oil_service(self,*args, **kwargs):
        if "oil_service_id" in kwargs and kwargs["oil_service_id"] is not None:
            return self.objects.filter(pk=kwargs['oil_service_id']).first()  
        if "oil_service_code" in kwargs and kwargs["oil_service_code"] is not None:
            return self.objects.filter(oil_service_code=kwargs['oil_service_code']).first()  
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        
        
    def add_oil_service(self,*args,**kwargs):
        result,message,oil_service=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_oil_service"):
            message="دسترسی غیر مجاز"
            return result,message,oil_service

        oil_service=OilService()
        if 'title' in kwargs:
            oil_service.title=kwargs["title"]
            if len(OilService.objects.filter(title=oil_service.title))>0:
                message='نام تکراری برای وسیله نقلیه جدید'
                return FAILED,message,None
        if 'owner_id' in kwargs:
            oil_service.owner_id=kwargs["owner_id"]
        if 'brand_name' in kwargs:
            oil_service.brand_name=kwargs["brand_name"]
        if 'model_name' in kwargs:
            oil_service.model_name=kwargs["model_name"]
        if 'oil_service_type' in kwargs:
            oil_service.oil_service_type=kwargs["oil_service_type"]
        if 'oil_service_color' in kwargs:
            oil_service.oil_service_color=kwargs["oil_service_color"]
        if 'oil_service_code' in kwargs:
            oil_service.oil_service_code=kwargs["oil_service_code"]
        if 'plaque' in kwargs:
            oil_service.plaque=kwargs["plaque"]
        if 'year' in kwargs:
            oil_service.year=kwargs["year"]
        if 'kilometer' in kwargs:
            oil_service.kilometer=kwargs["kilometer"]
        if 'driver_id' in kwargs:
            driver_id=kwargs["driver_id"]
            if driver_id is not None and driver_id>0:
                driver=DriverRepo(request=self.request).driver(driver_id=driver_id)
                if driver is not None:
                    oil_service.driver=driver.person_account.person.full_name
          
        if 'price' in kwargs:
            oil_service.price=kwargs["price"]
        
        if 'description' in kwargs:
            oil_service.description=kwargs["description"]
                
        (result,message,oil_service)=oil_service.save()
        return result,message,oil_service


class FilterServiceRepo():
    def __init__(self,request,*args, **kwargs):
        self.me=None
        self.my_accounts=[]
        self.request=request
        self.objects=FilterService.objects.filter(id=0)
        profile=PersonRepo(request=request).me
        if profile is not None:
            if request.user.has_perm(APP_NAME+".view_filter_service"):
                self.objects=FilterService.objects
                self.my_accounts=self.objects 
    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            objects=objects.filter(Q(name__contains=search_for) | Q(code=search_for)  )
        if "vehicle_id" in kwargs:
            vehicle_id=kwargs["vehicle_id"]
            objects=objects.filter(work_shift__vehicle_id=vehicle_id)  
        if "owner_id" in kwargs:
            owner_id=kwargs["owner_id"]
            objects=objects.filter(owner_id=owner_id)  
        return objects.all()
        
    def filter_service(self,*args, **kwargs):
        if "filter_service_id" in kwargs and kwargs["filter_service_id"] is not None:
            return self.objects.filter(pk=kwargs['filter_service_id']).first()  
        if "filter_service_code" in kwargs and kwargs["filter_service_code"] is not None:
            return self.objects.filter(filter_service_code=kwargs['filter_service_code']).first()  
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        
        
    def add_filter_service(self,*args,**kwargs):
        result,message,filter_service=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_filter_service"):
            message="دسترسی غیر مجاز"
            return result,message,filter_service

        filter_service=FilterService()
        if 'title' in kwargs:
            filter_service.title=kwargs["title"]
            if len(FilterService.objects.filter(title=filter_service.title))>0:
                message='نام تکراری برای وسیله نقلیه جدید'
                return FAILED,message,None
        if 'owner_id' in kwargs:
            filter_service.owner_id=kwargs["owner_id"]
        if 'brand_name' in kwargs:
            filter_service.brand_name=kwargs["brand_name"]
        if 'model_name' in kwargs:
            filter_service.model_name=kwargs["model_name"]
        if 'filter_service_type' in kwargs:
            filter_service.filter_service_type=kwargs["filter_service_type"]
        if 'filter_service_color' in kwargs:
            filter_service.filter_service_color=kwargs["filter_service_color"]
        if 'filter_service_code' in kwargs:
            filter_service.filter_service_code=kwargs["filter_service_code"]
        if 'plaque' in kwargs:
            filter_service.plaque=kwargs["plaque"]
        if 'year' in kwargs:
            filter_service.year=kwargs["year"]
        if 'kilometer' in kwargs:
            filter_service.kilometer=kwargs["kilometer"]
        if 'driver_id' in kwargs:
            driver_id=kwargs["driver_id"]
            if driver_id is not None and driver_id>0:
                driver=DriverRepo(request=self.request).driver(driver_id=driver_id)
                if driver is not None:
                    filter_service.driver=driver.person_account.person.full_name
        
        if 'price' in kwargs:
            filter_service.price=kwargs["price"]
        
        if 'description' in kwargs:
            filter_service.description=kwargs["description"]
                
        (result,message,filter_service)=filter_service.save()
        return result,message,filter_service
  
    
class ProductRepo():
    def __init__(self,request,*args, **kwargs):
        self.me=None
        self.my_accounts=[]
        self.request=request
        self.objects=Product.objects.filter(id=0)
        profile=PersonRepo(request=request).me
        if profile is not None:
            if request.user.has_perm(APP_NAME+".view_product"):
                self.objects=Product.objects
                self.my_accounts=self.objects 
    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            objects=objects.filter(Q(name__contains=search_for) | Q(code=search_for)  )
        if "vehicle_id" in kwargs:
            vehicle_id=kwargs["vehicle_id"]
            objects=objects.filter(work_shift__vehicle_id=vehicle_id)  
        if "owner_id" in kwargs:
            owner_id=kwargs["owner_id"]
            objects=objects.filter(owner_id=owner_id)  
        return objects.all()
        
    def product(self,*args, **kwargs):
        if "product_id" in kwargs and kwargs["product_id"] is not None:
            return self.objects.filter(pk=kwargs['product_id']).first()  
        if "product_code" in kwargs and kwargs["product_code"] is not None:
            return self.objects.filter(product_code=kwargs['product_code']).first()  
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        
        
    def add_product(self,*args,**kwargs):
        result,message,product=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_product"):
            message="دسترسی غیر مجاز"
            return result,message,product

        product=Product()
        if 'title' in kwargs:
            product.title=kwargs["title"]
            if len(Product.objects.filter(title=product.title))>0:
                message='نام تکراری برای وسیله نقلیه جدید'
                return FAILED,message,None
        if 'owner_id' in kwargs:
            product.owner_id=kwargs["owner_id"]
        if 'brand_name' in kwargs:
            product.brand_name=kwargs["brand_name"]
        if 'model_name' in kwargs:
            product.model_name=kwargs["model_name"]
        if 'product_type' in kwargs:
            product.product_type=kwargs["product_type"]
        if 'product_color' in kwargs:
            product.product_color=kwargs["product_color"]
        if 'product_code' in kwargs:
            product.product_code=kwargs["product_code"]
        if 'plaque' in kwargs:
            product.plaque=kwargs["plaque"]
        if 'year' in kwargs:
            product.year=kwargs["year"]
        if 'kilometer' in kwargs:
            product.kilometer=kwargs["kilometer"]
        if 'driver_id' in kwargs:
            driver_id=kwargs["driver_id"]
            if driver_id is not None and driver_id>0:
                driver=DriverRepo(request=self.request).driver(driver_id=driver_id)
                if driver is not None:
                    product.driver=driver.person_account.person.full_name
          
        if 'price' in kwargs:
            product.price=kwargs["price"]
        
        if 'description' in kwargs:
            product.description=kwargs["description"]
                
        (result,message,product)=product.save()
        return result,message,product
 
    
class VehicleRepo():
    def __init__(self,request,*args, **kwargs):
        self.me=None
        self.my_accounts=[]
        self.request=request
        self.objects=Vehicle.objects.filter(id=0)
        profile=PersonRepo(request=request).me
        if profile is not None:
            if request.user.has_perm(APP_NAME+".view_vehicle"):
                self.objects=Vehicle.objects
                self.my_accounts=self.objects 
    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            objects=objects.filter(Q(name__contains=search_for) | Q(code=search_for)  )
        if "parent_id" in kwargs:
            parent_id=kwargs["parent_id"]
            objects=objects.filter(parent_id=parent_id)  
        if "owner_id" in kwargs:
            owner_id=kwargs["owner_id"]
            objects=objects.filter(owner_id=owner_id)  
        return objects.all()
        
    def vehicle(self,*args, **kwargs):
        if "vehicle_id" in kwargs and kwargs["vehicle_id"] is not None:
            return self.objects.filter(pk=kwargs['vehicle_id']).first()  
        if "vehicle_code" in kwargs and kwargs["vehicle_code"] is not None:
            return self.objects.filter(vehicle_code=kwargs['vehicle_code']).first()  
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        
        
    def add_vehicle(self,*args,**kwargs):
        result,message,vehicle=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_vehicle"):
            message="دسترسی غیر مجاز"
            return result,message,vehicle

        vehicle=Vehicle()
        if 'title' in kwargs:
            vehicle.title=kwargs["title"]
            if len(Vehicle.objects.filter(title=vehicle.title))>0:
                message='نام تکراری برای وسیله نقلیه جدید'
                return FAILED,message,None
        if 'owner_id' in kwargs:
            vehicle.owner_id=kwargs["owner_id"]
        if 'brand_name' in kwargs:
            vehicle.brand_name=kwargs["brand_name"]
        if 'model_name' in kwargs:
            vehicle.model_name=kwargs["model_name"]
        if 'vehicle_type' in kwargs and kwargs['vehicle_type']:
            vehicle.vehicle_type=kwargs["vehicle_type"]
        if 'vehicle_color' in kwargs and kwargs['vehicle_color']:
            vehicle.vehicle_color=kwargs["vehicle_color"]
        if 'vehicle_size' in kwargs and kwargs['vehicle_size']:
            vehicle.vehicle_size=kwargs["vehicle_size"]
        if 'vehicle_code' in kwargs:  
            vehicle.vehicle_code=kwargs["vehicle_code"]
            if len(Vehicle.objects.filter(vehicle_code=vehicle.vehicle_code))>0:
                message='کد تکراری برای وسیله نقلیه جدید'
                return FAILED,message,None
        if 'plaque' in kwargs:
            vehicle.plaque=kwargs["plaque"]
        if 'year' in kwargs:
            vehicle.year=kwargs["year"]
        if 'kilometer' in kwargs:
            vehicle.kilometer=kwargs["kilometer"]
        if 'driver_id' in kwargs:
            driver_id=kwargs["driver_id"]
            if driver_id is not None and driver_id>0:
                driver=DriverRepo(request=self.request).driver(driver_id=driver_id)
                if driver is not None:
                    vehicle.driver=driver.person_account.person.full_name
          
        if 'price' in kwargs:
            vehicle.price=kwargs["price"]
        
        if 'description' in kwargs:
            vehicle.description=kwargs["description"]
                
        (result,message,vehicle)=vehicle.save()
        return result,message,vehicle

    def import_vehicles_from_excel(self,*args,**kwargs):
        result,message,vehicles=FAILED,"",[]
        excel_file=kwargs['excel_file']
         
        import openpyxl 

        wb = openpyxl.load_workbook(excel_file)
        try:
            ws = wb['vehicles']
            

        except:
            message='فایل شما برگه ماشین آلات ندارد.'
            return result,message,None
        count=kwargs['count']
        try:
            count=int(ws.cell(row=1, column=2).value)
        except:
            
            message='فایل برگه ماشین آلات ، تعداد ندارد.'
            return result,message,None 

        vehicles_to_import=[]

        START_ROW=EXCEL_VEHICLES_DATA_START_ROW
        for i in range(START_ROW,count+START_ROW):
            vehicle={}
            
            i=str(i) 
            # product['id']=ws['A'+str(i)].value
            iiiddd=ws['B'+i].value
            if iiiddd is not None:
                id=0
                try:
                    id=int(ws['B'+i].value)
                except:
                    pass

                vehicle_code=(ws['C'+i].value)
                vehicle_type=(ws['D'+i].value)
                title=(ws['E'+i].value)
                year=(ws['F'+i].value)
                plaque=(ws['G'+i].value)
                color=(ws['H'+i].value)
                kilometer=0
                
                try:
                    kilometer=int(ws['I'+i].value)
                except:
                    pass

                vehicle['id']=id
                description=(ws['J'+i].value)
                vehicle['vehicle_code']=vehicle_code
                vehicle['vehicle_type']=vehicle_type
                vehicle['title']=title
                vehicle['year']=year
                vehicle['plaque']=plaque
                vehicle['color']=color
                vehicle['kilometer']=kilometer
                vehicle['description']=description
                # vehicle['thumbnail_origin']=ws['F'+str(i)].value
                if vehicle['title'] is not None and not vehicle['title']=="":
                    if vehicle['vehicle_code'] is not None and not vehicle['vehicle_code']=="":
                        vehicles_to_import.append(vehicle) 
        modified=added=0

        for vehicle in vehicles_to_import:
            old_vehicle=Vehicle.objects.filter(vehicle_code=vehicle["vehicle_code"]).first()
            if old_vehicle is not None:
                old_vehicle.vehicle_code=vehicle["vehicle_code"]
                old_vehicle.vehicle_type=vehicle["vehicle_type"]
                old_vehicle.title=vehicle["title"]
                old_vehicle.year=vehicle["year"]
                old_vehicle.plaque=vehicle["plaque"]
                old_vehicle.color=vehicle["color"]
                old_vehicle.kilometer=vehicle["kilometer"]
                old_vehicle.description=vehicle["description"]
                old_vehicle.save()
                modified+=1
            else:
                result,message,new_vehicle=self.add_vehicle(**vehicle)
          
                if result==SUCCEED:
                    added+=1
        result=SUCCEED
        message=f"""{added} سرویس اضافه شد.
                    <br>
                    {modified} سرویس ویرایش شد. """
        services=self.list()

        
        if True:
            log_data={}
            from log.repo import LogRepo
            log_data['person_id']=PersonRepo(request=self.request).me.id
            log_data['url']=reverse("accounting:services")
            log_data['title']="بازیابی سرویس ها"
            log_data['description']=message
            log_data['app_name']=APP_NAME
            LogRepo(request=self.request).add_log(**log_data)

        return result,message,services


class DriverRepo():
    def __init__(self,request,*args, **kwargs):
        self.me=None
        self.my_accounts=[]
        self.request=request
        self.objects=Driver.objects.filter(id=0)
        profile=PersonRepo(request=request).me
        if profile is not None:
            if request.user.has_perm(APP_NAME+".view_vehicle"):
                self.objects=Driver.objects
                self.my_accounts=self.objects 
    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            objects=objects.filter(Q(name__contains=search_for) | Q(code=search_for)  )
        if "parent_id" in kwargs:
            parent_id=kwargs["parent_id"]
            objects=objects.filter(parent_id=parent_id)  
        if "owner_id" in kwargs:
            owner_id=kwargs["owner_id"]
            objects=objects.filter(owner_id=owner_id)  
        return objects.all()
        
    def driver(self,*args, **kwargs):
        if "driver_id" in kwargs and kwargs["driver_id"] is not None:
            return self.objects.filter(pk=kwargs['driver_id']).first()  
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        
        
    def add_driver(self,*args,**kwargs):
        result,message,driver=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_driver"):
            message="دسترسی غیر مجاز"
            return result,message,driver

        driver=Driver()
        if 'title' in kwargs:
            driver.title=kwargs["title"]
            if len(Driver.objects.filter(title=driver.title))>0:
                message='نام تکراری برای وسیله نقلیه جدید'
                return FAILED,message,None
        if 'owner_id' in kwargs:
            driver.owner_id=kwargs["owner_id"]
        if 'brand_name' in kwargs:
            driver.brand_name=kwargs["brand_name"]
        if 'model_name' in kwargs:
            driver.model_name=kwargs["model_name"]
        if 'driver_type' in kwargs:
            driver.driver_type=kwargs["driver_type"]
        if 'driver_color' in kwargs:
            driver.driver_color=kwargs["driver_color"]
        if 'driver_code' in kwargs:
            driver.driver_code=kwargs["driver_code"]
        if 'plaque' in kwargs:
            driver.plaque=kwargs["plaque"]
        if 'year' in kwargs:
            driver.year=kwargs["year"]
        if 'kilometer' in kwargs:
            driver.kilometer=kwargs["kilometer"]
        if 'driver_id' in kwargs:
            if kwargs['driver_id']>0:
                driver.driver_id=kwargs["driver_id"]
          
        (result,message,driver)=driver.save()
        return result,message,driver


class MaintenanceRepo():
    def __init__(self,request,*args, **kwargs):
        self.me=None
        self.my_accounts=[]
        self.request=request
        self.objects=Maintenance.objects.filter(id=0)
        profile=PersonRepo(request=request).me
        if profile is not None:
            if request.user.has_perm(APP_NAME+".view_maintenance"):
                self.objects=Maintenance.objects
                self.my_accounts=self.objects 
    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            objects=objects.filter(Q(name__contains=search_for) | Q(code=search_for)  )
        if "parent_id" in kwargs:
            parent_id=kwargs["parent_id"]
            objects=objects.filter(parent_id=parent_id)  
        if "vehicle_id" in kwargs:
            vehicle_id=kwargs["vehicle_id"]
            objects=objects.filter(vehicle_id=vehicle_id)  
        if "service_man_id" in kwargs:
            service_man_id=kwargs["service_man_id"]
            objects=objects.filter(service_man_id=service_man_id)
        return objects.all()
        
    def maintenance(self,*args, **kwargs):
        if "maintenance_id" in kwargs and kwargs["maintenance_id"] is not None:
            return self.objects.filter(pk=kwargs['maintenance_id']).first()  
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        
    
       

    def add_invoice(self,*args,**kwargs):
        result,message,invoice=FAILED,"",None 
        from accounting.models import Invoice,PersianCalendar

            
        if not self.request.user.has_perm(APP_NAME+".add_invoice"):
            message="دسترسی غیر مجاز"
            return result,message,invoice

        invoice=Invoice()
        
        if 'valid' in kwargs and kwargs['valid'] is not None:
            invoice.valid=kwargs["valid"]

        if 'title' in kwargs:
            invoice.title=kwargs["title"]
        if 'parent_id' in kwargs:
            if kwargs["parent_id"]>0:
                invoice.parent_id=kwargs["parent_id"]
        if 'color' in kwargs:
            invoice.color=kwargs["color"]
        if 'code' in kwargs:
            invoice.code=kwargs["code"]
        if 'priority' in kwargs:
            invoice.priority=kwargs["priority"]
        if 'bedehkar_id' in kwargs:
            invoice.bedehkar_id=kwargs["bedehkar_id"]
        if 'bestankar_id' in kwargs:
            invoice.bestankar_id=kwargs["bestankar_id"]
        if 'event_datetime' in kwargs:
            
            year=kwargs['event_datetime'][:2]
            if year=="13" or year=="14":
                kwargs['event_datetime']=PersianCalendar().to_gregorian(kwargs["event_datetime"])
            invoice.event_datetime=kwargs["event_datetime"]

        if 'type' in kwargs:
            invoice.type=kwargs["type"]

           
        if 'status' in kwargs:
            invoice.status=kwargs["status"]

           
           
        if 'invoice_no' in kwargs:
            invoice.invoice_no=kwargs["invoice_no"]


        if 'maintenance_id' in kwargs:
            maintenance_id=kwargs["maintenance_id"]
            maintenance=Maintenance.objects.filter(pk=maintenance_id).first()
            if maintenance is not None:
                (result,message,invoice)=invoice.save()
                maintenance.invoices.add(invoice.id)
                result=SUCCEED
                message='با موفقیت اضافه شد.'

        return result,message,invoice

     
    def add_invoice_to_maintenance(self,*args, **kwargs):   
        result,message,invoice=FAILED,'',None
        if not self.request.user.has_perm(APP_NAME+".change_maintenance"):
            message="دسترسی غیر مجاز"
            return result,message,invoice
        maintenance=Maintenance.objects.filter(pk=kwargs['maintenance_id']).first()
        from accounting.repo import InvoiceRepo
        invoice=InvoiceRepo(request=self.request).invoice(pk=kwargs['invoice_id'])
        if maintenance is not None and invoice is not None:
            maintenance.invoices.add(invoice.id)
            result=SUCCEED
            message='با موفقیت اضافه شد.'
        return result,message,invoice
    
    def add_maintenance(self,*args,**kwargs):
        result,message,maintenance=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_maintenance"):
            message="دسترسی غیر مجاز"
            return result,message,maintenance

        maintenance=Maintenance()
        if 'title' in kwargs:
            maintenance.title=kwargs["title"]

        if 'hour' in kwargs:
            maintenance.hour=kwargs["hour"]
            
        if 'service_man_id' in kwargs:
            maintenance.service_man_id=kwargs["service_man_id"]
            
        if 'vehicle_id' in kwargs:
            maintenance.vehicle_id=kwargs["vehicle_id"]
            
        if 'driver_id' in kwargs:
            maintenance.driver_id=kwargs["driver_id"]

        if 'kilometer' in kwargs:
            maintenance.kilometer=kwargs["kilometer"]

        if 'maintenance_type' in kwargs:
            maintenance.maintenance_type=kwargs["maintenance_type"]

        if 'description' in kwargs:
            maintenance.description=kwargs["description"]


        if 'event_datetime' in kwargs and kwargs['event_datetime'] is not None and not kwargs['event_datetime']=='':
            year=kwargs['event_datetime'][:2]
            if year=="13" or year=="14":
                kwargs['event_datetime']=PersianCalendar().to_gregorian(kwargs["event_datetime"])
            maintenance.event_datetime=kwargs["event_datetime"]
         
        (result,message,maintenance)=maintenance.save()
        return result,message,maintenance

 
class ServiceManRepo():
    def __init__(self,request,*args, **kwargs):
        self.me=None
        self.my_accounts=[]
        self.request=request
        self.objects=ServiceMan.objects.filter(id=0)
        profile=PersonRepo(request=request).me
        if profile is not None:
            if request.user.has_perm(APP_NAME+".view_vehicle"):
                self.objects=ServiceMan.objects
                self.my_accounts=self.objects 
    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            objects=objects.filter(Q(person_account__person__full_name__contains=search_for)    )
        if "parent_id" in kwargs:
            parent_id=kwargs["parent_id"]
            objects=objects.filter(parent_id=parent_id)  
        return objects.all()
        
    def service_man(self,*args, **kwargs):
        if "service_man_id" in kwargs and kwargs["service_man_id"] is not None:
            return self.objects.filter(pk=kwargs['service_man_id']).first()  
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        
        
    def add_service_man(self,*args,**kwargs):
        result,message,service_man=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_service_man"):
            message="دسترسی غیر مجاز"
            return result,message,service_man
        if len(ServiceMan.objects.filter(person_account_id=kwargs["person_account_id"]))>0:
            message='قبلا برای این شخص سرویس کار ایجاد شده است.'
            return FAILED,message,None
        service_man=ServiceMan() 
        if 'person_account_id' in kwargs:
            service_man.person_account_id=kwargs["person_account_id"]
          
        (result,message,service_man)=service_man.save()
        return result,message,service_man

  
class VehicleEventRepo():
    def __init__(self,request,*args, **kwargs):
        self.me=None
        self.my_accounts=[]
        self.request=request
        self.objects=VehicleEvent.objects.filter(id=0)
        profile=PersonRepo(request=request).me
        if profile is not None:
            if request.user.has_perm(APP_NAME+".view_vehicle"):
                self.objects=VehicleEvent.objects
                self.my_accounts=self.objects 
    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            objects=objects.filter(Q(person_account__person__full_name__contains=search_for)    )
        if "vehicle_id" in kwargs:
            vehicle_id=kwargs["vehicle_id"]
            objects=objects.filter(vehicle_id=vehicle_id)  
        return objects.all()
        
    def vehicle_event(self,*args, **kwargs):
        if "vehicle_event_id" in kwargs and kwargs["vehicle_event_id"] is not None:
            return self.objects.filter(pk=kwargs['vehicle_event_id']).first()  
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        
        
    def add_vehicle_event(self,*args,**kwargs):
        result,message,vehicle_event=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_vehicle_event"):
            message="دسترسی غیر مجاز"
            return result,message,vehicle_event
        if len(VehicleEvent.objects.filter(person_account_id=kwargs["person_account_id"]))>0:
            message='قبلا برای این شخص سرویس کار ایجاد شده است.'
            return FAILED,message,None
        vehicle_event=VehicleEvent() 
        if 'person_account_id' in kwargs:
            vehicle_event.person_account_id=kwargs["person_account_id"]
          
        (result,message,vehicle_event)=vehicle_event.save()
        return result,message,vehicle_event


class TavaghofRepo():
    def __init__(self,request,*args, **kwargs):
        self.me=None
        self.my_accounts=[]
        self.request=request
        self.objects=Tavaghof.objects.filter(id=0)
        profile=PersonRepo(request=request).me
        if profile is not None:
            if request.user.has_perm(APP_NAME+".view_vehicle"):
                self.objects=Tavaghof.objects
                self.my_accounts=self.objects 
    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            objects=objects.filter(Q(person_account__person__full_name__contains=search_for)    )
        if "parent_id" in kwargs:
            parent_id=kwargs["parent_id"]
            objects=objects.filter(parent_id=parent_id)  
        if "vehicle_id" in kwargs:
                    vehicle_id=kwargs["vehicle_id"]
                    objects=objects.filter(vehicle_id=vehicle_id) 
        return objects.all()
        
    def tavaghof(self,*args, **kwargs):
        if "tavaghof_id" in kwargs and kwargs["tavaghof_id"] is not None:
            return self.objects.filter(pk=kwargs['tavaghof_id']).first()  
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        
        
    def add_tavaghof(self,*args,**kwargs):
        result,message,tavaghof=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_tavaghof"):
            message="دسترسی غیر مجاز"
            return result,message,tavaghof
        if len(Tavaghof.objects.filter(person_account_id=kwargs["person_account_id"]))>0:
            message='قبلا برای این شخص سرویس کار ایجاد شده است.'
            return FAILED,message,None
        tavaghof=Tavaghof() 
        if 'person_account_id' in kwargs:
            tavaghof.person_account_id=kwargs["person_account_id"]
          
        (result,message,tavaghof)=tavaghof.save()
        return result,message,tavaghof

 
class ServiceRepo():
    def __init__(self,request,*args, **kwargs):
        self.me=None
        self.my_accounts=[]
        self.request=request
        self.objects=Service.objects.filter(id=0)
        profile=PersonRepo(request=request).me
        if profile is not None:
            if request.user.has_perm(APP_NAME+".view_vehicle"):
                self.objects=Service.objects
                self.my_accounts=self.objects 
    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            objects=objects.filter(Q(person_account__person__full_name__contains=search_for)    )
        if "driver_id" in kwargs:
                    driver_id=kwargs["driver_id"]
                    objects=objects.filter(driver_id=driver_id) 

        if "service_man_id" in kwargs:
                    service_man_id=kwargs["service_man_id"]
                    objects=objects.filter(service_man_id=service_man_id)  
        if "vehicle_id" in kwargs:
            vehicle_id=kwargs["vehicle_id"]
            objects=objects.filter(vehicle_id=vehicle_id) 

        if "shift" in kwargs and kwargs['shift']: 
            objects=objects.filter(shift=kwargs["shift"]) 

        if "from_shift_date" in kwargs and kwargs['from_shift_date']:
            year=kwargs['from_shift_date'][:2]
            if year=="13" or year=="14":
                kwargs['from_shift_date']=PersianCalendar().to_gregorian(kwargs["from_shift_date"])
            
            objects=objects.filter(shift_date__gte=kwargs["from_shift_date"]) 

        if "to_shift_date" in kwargs and kwargs['to_shift_date']:
            year=kwargs['to_shift_date'][:2]
            if year=="13" or year=="14":
                kwargs['to_shift_date']=PersianCalendar().to_gregorian(kwargs["to_shift_date"])

            import datetime 
            delta=datetime.timedelta(hours=23,minutes=59,seconds=59)
            kwargs['to_shift_date']=kwargs['to_shift_date']+delta                        
            objects=objects.filter(shift_date__lte=kwargs["to_shift_date"]) 


        return objects.all()
        
    def service(self,*args, **kwargs):
        if "service_id" in kwargs and kwargs["service_id"] is not None:
            return self.objects.filter(pk=kwargs['service_id']).first()  
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        
        
    def add_service(self,*args,**kwargs):
        result,message,service=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_service"):
            message="دسترسی غیر مجاز"
            return result,message,service
         
        service=Service() 
        if 'grease' in kwargs:
            service.grease=kwargs["grease"]
          
        if 'oil_liter' in kwargs:
            service.oil_liter=kwargs["oil_liter"]
          
        if 'filter_type' in kwargs:
            service.filter_type=kwargs["filter_type"]
          
        if 'oil_type' in kwargs:
            service.oil_type=kwargs["oil_type"]
          
        if 'filter_action' in kwargs:
            service.filter_action=kwargs["filter_action"]
          
        if 'driver_id' in kwargs and kwargs["driver_id"]:
            service.driver_id=kwargs["driver_id"]
          
        if 'vehicle_id' in kwargs and kwargs["vehicle_id"]:
            service.vehicle_id=kwargs["vehicle_id"]
          
        if 'service_man_id' in kwargs and kwargs["service_man_id"]:
            service.service_man_id=kwargs["service_man_id"]
          
        if 'shift' in kwargs and kwargs["shift"]:
            service.shift=kwargs["shift"]
          
        if 'description' in kwargs and kwargs["description"]:
            service.description=kwargs["description"]
          
        if 'vehicle_hour' in kwargs and kwargs["vehicle_hour"]:
            service.vehicle_hour=kwargs["vehicle_hour"]
          
        if "shift_date" in kwargs and kwargs['shift_date']:
            year=kwargs['shift_date'][:2]
            if year=="13" or year=="14":
                kwargs['shift_date']=PersianCalendar().to_gregorian(kwargs["shift_date"])
            service.shift_date=kwargs["shift_date"]
        (result,message,service)=service.save()
        return result,message,service

 
class AnbarProductRepo():
    def __init__(self,request,*args, **kwargs):
        self.me=None
        self.my_accounts=[]
        self.request=request
        self.objects=AnbarProduct.objects.filter(id=0)
        profile=PersonRepo(request=request).me
        if profile is not None:
            if request.user.has_perm(APP_NAME+".view_vehicle"):
                self.objects=AnbarProduct.objects
                self.my_accounts=self.objects 
    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            objects=objects.filter(Q(person_account__person__full_name__contains=search_for)    )
        if "parent_id" in kwargs:
            parent_id=kwargs["parent_id"]
            objects=objects.filter(parent_id=parent_id)  
         

        if "vehicle_code" in kwargs and kwargs['vehicle_code']:
            objects=objects.filter(vehicle__vehicle_code=kwargs["vehicle_code"])  
        if "driver_id" in kwargs and kwargs["driver_id"]:
            pass
            # objects=objects.filter(driver_id=kwargs["driver_id"]) 
        return objects.all()
        
    def anbar_product(self,*args, **kwargs):
        if "anbar_product_id" in kwargs and kwargs["anbar_product_id"] is not None:
            return self.objects.filter(pk=kwargs['anbar_product_id']).first()  
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        
        
    def add_anbar_product(self,*args,**kwargs):
        result,message,anbar_product=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_anbar_product"):
            message="دسترسی غیر مجاز"
            return result,message,anbar_product
        if len(AnbarProduct.objects.filter(person_account_id=kwargs["person_account_id"]))>0:
            message='قبلا برای این شخص سرویس کار ایجاد شده است.'
            return FAILED,message,None
        anbar_product=AnbarProduct() 
        if 'person_account_id' in kwargs:
            anbar_product.person_account_id=kwargs["person_account_id"]
        
        (result,message,anbar_product)=anbar_product.save()
        return result,message,anbar_product
  