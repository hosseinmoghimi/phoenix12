from .models import Shop,ShopPackage,Supplier,Customer,CartItem,Shipper,CustomerGroup

from .apps import APP_NAME
from .enums import *
from log.repo import LogRepo 
from django.db.models import Q
from django.shortcuts import reverse
from authentication.repo import PersonRepo
from accounting.repo import InvoiceLineItemUnitRepo
from utility.num import filter_number
from utility.calendar import PersianCalendar,to_gregorian
from utility.constants import FAILED,SUCCEED
from utility.log import leolog
from authentication.repo import PersonRepo
from accounting.repo import PersonCategoryEnum
from .constants import EXCEL_SHOPS_DATA_START_ROW

class ShopPackageRepo():
    def __init__(self,request,*args, **kwargs):
        self.me=None
        self.my_accounts=[]
        self.request=request
        self.objects=ShopPackage.objects.filter(id=0)
        profile=PersonRepo(request=request).me
        if profile is not None:
            if request.user.has_perm(APP_NAME+".view_shoppackage"):
                self.objects=ShopPackage.objects
                self.my_accounts=self.objects 
     
    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            objects=objects.filter(Q(name__contains=search_for) | Q(code=search_for)  )
        if "parent_id" in kwargs:
            parent_id=kwargs["parent_id"]
            objects=objects.filter(parent_id=parent_id)  
        if "supplier_id" in kwargs:
            supplier_id=kwargs["supplier_id"]
            objects=objects.filter(supplier_id=supplier_id)
        if "product_id" in kwargs:
            product_id=kwargs["product_id"]
            objects=objects.filter(product_id=product_id)
          
        return objects.all()
        
    def shop_package(self,*args, **kwargs):
        if "shop_package_id" in kwargs and kwargs["shop_package_id"] is not None:
            return self.objects.filter(pk=kwargs['shop_package_id']).first()  
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        
        
    def add_shop_package(self,*args,**kwargs):
        result,message,shop_package=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_shop_package"):
            message="دسترسی غیر مجاز"
            return result,message,shop_package

        shop_package=ShopPackage()
        if 'title' in kwargs:
            shop_package.title=kwargs["title"]
        if 'parent_id' in kwargs:
            if kwargs["parent_id"]>0:
                shop_package.parent_id=kwargs["parent_id"]
        if 'color' in kwargs:
            shop_package.color=kwargs["color"]
        if 'supplier_id' in kwargs:
            shop_package.supplier_id=kwargs["supplier_id"]
        if 'priority' in kwargs:
            shop_package.priority=kwargs["priority"]
        if 'type' in kwargs:
            shop_package.type=kwargs["type"]

         
        (result,message,shop_package)=shop_package.save()
        return result,message,shop_package


class ShopRepo():
    def __init__(self,request,*args, **kwargs):
        self.me=None
        self.my_accounts=[]
        self.request=request
        self.objects=Shop.objects.filter(id=0)
        me_person=PersonRepo(request=request).me
        me_customer=CustomerRepo(request=request).me
        me_supplier=SupplierRepo(request=request).me
        if request.user.has_perm(APP_NAME+".view_shop"):
            self.objects=Shop.objects 
        elif me_customer is not None:
            regions_ids=[]
            for region in me_customer.all_regions():
                regions_ids.append(region.id)
            groups_ids=[]
            for group in me_customer.groups.all():
                groups_ids.append(group.id)
            self.objects=Shop.objects.filter(group_id__in=groups_ids).filter(region_id__in=regions_ids)
        elif me_supplier is not None: 
            self.objects=Shop.objects.filter(supplier_id=me_supplier.id)
        
    def primary_shop(self,product,*args, **kwargs):
        if product is None :
            return None
        shops=self.objects.filter(product_id=product.id) 
        if len(shops)>0:
            return shops.order_by('unit_price').first()
        # return Shop(unit_name=UnitNameEnum.ADAD,unit_price=0)
        return None
    
    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            objects=objects.filter(Q(name__contains=search_for) | Q(code=search_for)  )
        if "parent_id" in kwargs:
            parent_id=kwargs["parent_id"]
            objects=objects.filter(parent_id=parent_id)  
        if "supplier_id" in kwargs:
            supplier_id=kwargs["supplier_id"]
            objects=objects.filter(supplier_id=supplier_id)
        if "product_id" in kwargs:
            product_id=kwargs["product_id"]
            objects=objects.filter(product_id=product_id)
        if "region_id" in kwargs:
            region_id=kwargs["region_id"]
            objects=objects.filter(region_id=region_id)
        if "group_id" in kwargs:
            group_id=kwargs["group_id"]
            objects=objects.filter(group_id=group_id)
          
        return objects.all()
        
    def shop(self,*args, **kwargs):
        if "shop_id" in kwargs and kwargs["shop_id"] is not None:
            return self.objects.filter(pk=kwargs['shop_id']).first()  
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        
    def add_shop(self,*args,**kwargs):
        result,message,shop=FAILED,"",None
        
        me_supplier=SupplierRepo(request=self.request).me

        if me_supplier is None:
            if self.request.user.has_perm(APP_NAME+".add_shop"):
                if 'supplier_id' in kwargs:
                    supplier_id=kwargs["supplier_id"]
            else:
                message="دسترسی غیر مجاز"
                return result,message,shop
        else:
            supplier_id=me_supplier.id
        shop=Shop(supplier_id=supplier_id)
         
        if 'unit_price' in kwargs:
            if kwargs["unit_price"]>0:
                shop.unit_price=kwargs["unit_price"]
        if 'unit_name' in kwargs:
            shop.unit_name=kwargs["unit_name"]
         
        if 'available' in kwargs:
            shop.available=kwargs["available"]
            shop.quantity=kwargs["available"]
         
        if 'region_id' in kwargs:
            shop.region_id=kwargs["region_id"]
        if 'group_id' in kwargs:
            shop.group_id=kwargs["group_id"]
        if 'product_id' in kwargs:
            shop.product_id=kwargs["product_id"]
        if 'start_date' in kwargs:
            start_date=kwargs["start_date"]
            shop.start_date=to_gregorian(start_date)
        if 'end_date' in kwargs:
            end_date=kwargs["end_date"]
            shop.end_date=to_gregorian(end_date)



        if 'discount_percentage' in kwargs:
            shop.discount_percentage=kwargs["discount_percentage"]
        
        
         
        (result,message,shop)=shop.save()
        return result,message,shop

    def import_shops_from_excel(self,*args,**kwargs):
        result,message,shops=FAILED,"",[]
        excel_file=kwargs['excel_file']
        # import pandas
        
        # df = pandas.read_excel(excel_file)
        # shops=[]
        # for row in df.columns[0]:
        #     print (df.columns)
        import openpyxl 

        wb = openpyxl.load_workbook(excel_file)
        try:
            ws = wb['shops']
        
        except:
            message='فایل شما برگه آماده فروش ها ندارد.'
            return result,message,None
        count=kwargs['count']
        try:
            count=int(ws.cell(row=1, column=2).value)
        except:
            
            message='فایل برگه حساب ها ، تعداد ندارد.'
            return result,message,None  
        shops_to_import=[]
        START_ROW=EXCEL_SHOPS_DATA_START_ROW
        for i in range(START_ROW,count+START_ROW):
            shop={}
            
            
            i=str(i) 
            # product['id']=ws['A'+str(i)].value
            iiiddd=ws['A'+i].value
            if iiiddd is not None:
                id=int(ws['B'+i].value) 
                product_id=int(ws['E'+i].value)
                product_barcode=(ws['F'+i].value)
                supplier_id=int(ws['C'+i].value)
                group_id=int(ws['H'+i].value)
                region_id=int(ws['J'+i].value)
                quantity=int(ws['L'+i].value)
                unit_price=int(ws['O'+i].value)
                unit_name=(ws['M'+i].value)
                start_date=(ws['P'+i].value)
                end_date=(ws['Q'+i].value)
                discount_percentage=(ws['N'+i].value)

                shop['id']=0
                shop['product_id']=product_id
                shop['product_barcode']=product_barcode
                shop['supplier_id']=supplier_id
                shop['quantity']=quantity
                shop['group_id']=group_id
                shop['region_id']=region_id
                shop['unit_price']=unit_price
                shop['unit_name']=unit_name
                shop['start_date']=start_date
                shop['end_date']=end_date
                shop['discount_percentage']=discount_percentage
                shops_to_import.append(shop) 
        modified=added=0
        for shop in shops_to_import:
            from accounting.repo import ProductRepo
            from utility.repo import RegionRepo
            product=ProductRepo(request=self.request).product(barcode=shop["product_barcode"],product_id=shop['product_id'])
            supplier=SupplierRepo(request=self.request).supplier(supplier_id=shop["supplier_id"])
            region=RegionRepo(request=self.request).region(region_id=shop["region_id"])
            group=CustomerGroupRepo(request=self.request).customer_group(customer_group_id=shop["group_id"])
            Shop.objects.filter(supplier_id=supplier.id).filter(product_id=product.id).filter(region_id=region.id).filter(group_id=group.id).delete()
            result,message,new_shop=self.add_shop(
                product_id=product.id,
                supplier_id=supplier.id,
                group_id=group.id,
                region_id=region.id,
                unit_price=shop['unit_price'],
                unit_name=shop['unit_name'],
                quantity=shop['quantity'],
                available=shop['quantity'],
                discount_percentage=shop['discount_percentage'],
                start_date=PersianCalendar().to_gregorian(shop['start_date']),
                end_date=PersianCalendar().to_gregorian(shop['end_date']),
                                                        
            )
            shops.append(new_shop)
               
                # new_shop.title=shop["title"]
                # new_shop.barcode=shop["barcode"]
                # new_shop.unit_name=shop["unit_name"]
                # new_shop.unit_price=shop["unit_price"] 
                # new_shop.save()
            if result==SUCCEED:
                added+=1
        result=SUCCEED
        message=f"""{added} آماده فروش اضافه شد.
                    <br>
                    {modified} آماده فروش ویرایش شد. """
        shops=self.list()


        if True:
            log_data={}
            from log.repo import LogRepo
            log_data['person_id']=PersonRepo(request=self.request).me.id
            log_data['url']=reverse("market:shops")
            log_data['title']="بازیابی آماده فروش ها"
            log_data['description']=message
            log_data['app_name']=APP_NAME
            LogRepo(request=self.request).add_log(**log_data)
        
        return result,message,shops
    
class CustomerRepo():
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.objects=Customer.objects.filter(pk=0)
        person=PersonRepo(request=request).me
        self.me=Customer.objects.filter(person_account__person_id=person.id).first()
        if request.user.has_perm(APP_NAME+'.view_customer'):
            self.objects=Customer.objects
        elif person is not None:
            self.objects=Customer.objects.filter(person_account__person_id=person.id)


        
        if person is not None:
            self.me=Customer.objects.filter(person_account__person_id=person.id).first()

    def list(self,*args, **kwargs):
        objects=self.objects
        pure_code="876454453342236"
        try:
            pure_code=int(kwargs["search_for"]) 
        except:
            pass
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]

            objects=objects.filter(Q(name__contains=search_for) | Q(code=search_for) | Q(pure_code=pure_code ) )
        return objects.all()
     
    def customer(self,*args, **kwargs):
        if "customer_id" in kwargs and kwargs["customer_id"] is not None:
            return self.objects.filter(pk=kwargs['customer_id']).first()
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "code" in kwargs and kwargs["code"] is not None:
            return self.objects.filter(code=kwargs['code']).first()
             
        if "account_code" in kwargs and kwargs["account_code"] is not None:
            a= self.objects.filter(code=kwargs['account_code']).first() 
            if a is not None:
                return a
            else:
                try:
                    a= self.objects.filter(pure_code=filter_number(kwargs['account_code'])).first() 
                    if a is not None:
                        return a
                except:
                    pass
        
        
    def add_customer(self,*args,**kwargs):
        result,message,customer=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_customer") :
            message="دسترسی غیر مجاز"
            return result,message,customer
        

        customer=Customer() 
        if 'person_account_id' in kwargs and kwargs['person_account_id'] is not None and kwargs['person_account_id']>0:
            person_account_id=kwargs["person_account_id"]
            customer.person_account_id=person_account_id
            if len(Customer.objects.filter(person_account_id=person_account_id))>0:
                message='برای این حساب قبلا مشتری ایجاد شده است.'
                return result,message,None
         
 
        (result,message,customer)=customer.save() 
        if 'region_id' in kwargs:
            customer.regions.add(kwargs["region_id"])
        if 'groups_ids' in kwargs:
            groups_ids=(kwargs["groups_ids"])
            leolog(groups_ids=groups_ids)
            for group_id in groups_ids:
                group_id=int(group_id)
                leolog(group_id=group_id)
                customer.groups.add(group_id)
 
        return result,message,customer


class ShipperRepo():
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        self.objects=Shipper.objects
        me_person=PersonRepo(request=request).me
        if me_person is not None:
            self.me=self.objects.filter(person_account__person_id=me_person.id).first()
    def list(self,*args, **kwargs):
        objects=self.objects
        pure_code="876454453342236"
        try:
            pure_code=int(kwargs["search_for"]) 
        except:
            pass
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]

            objects=objects.filter(Q(name__contains=search_for) | Q(code=search_for) | Q(pure_code=pure_code ) )
        return objects.all()
     
    def shipper(self,*args, **kwargs):
        if "shipper_id" in kwargs and kwargs["shipper_id"] is not None:
            return self.objects.filter(pk=kwargs['shipper_id']).first()
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "code" in kwargs and kwargs["code"] is not None:
            return self.objects.filter(code=kwargs['code']).first()
             
        if "account_code" in kwargs and kwargs["account_code"] is not None:
            a= self.objects.filter(code=kwargs['account_code']).first() 
            if a is not None:
                return a
            else:
                try:
                    a= self.objects.filter(pure_code=filter_number(kwargs['account_code'])).first() 
                    if a is not None:
                        return a
                except:
                    pass
     
 
    def add_shipper(self,*args,**kwargs):
        result,message,shipper=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_shipper") :
            message="دسترسی غیر مجاز"
            return result,message,shipper
        

        shipper=Shipper() 
        if 'person_account_id' in kwargs and kwargs['person_account_id'] is not None and kwargs['person_account_id']>0:
            person_account_id=kwargs["person_account_id"]
            shipper.person_account_id=person_account_id
            if len(Shipper.objects.filter(person_account_id=person_account_id))>0:
                message='برای این حساب قبلا فروشنده ایجاد شده است.'
                return result,message,None
         
        (result,message,shipper)=shipper.save() 


        if 'region_id' in kwargs:
            shipper.regions.add(kwargs["region_id"])
        return result,message,shipper


class CartItemRepo():
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        # profile=PersonRepo(request=request).me
        self.objects=CartItem.objects
        # if profile is not None:
        #     self.me=self.objects.filter(profile=profile).first()
    def list(self,*args, **kwargs):
        objects=self.objects
        
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            objects=objects.filter(Q(name__contains=search_for) | Q(code=search_for) | Q(pure_code=pure_code ) )
            
        if "shop_id" in kwargs:
            shop_id=kwargs["shop_id"]
            objects=objects.filter(shop_id=shop_id)

            
        if "customer_id" in kwargs:
            customer_id=kwargs["customer_id"]
            objects=objects.filter(customer_id=customer_id)
            
        if "supplier_id" in kwargs:
            supplier_id=kwargs["supplier_id"]
            objects=objects.filter(shop__supplier_id=supplier_id)

        if "product_id" in kwargs:
            product_id=kwargs["product_id"]
            objects=objects.filter(shop__product_id=product_id)
   
        return objects.all()
   
    def add_cart_item(self,*args,**kwargs):
        result,message,cart_item,cart_items=FAILED,"",None,[]
        me_customer=CustomerRepo(request=self.request).me
        if  me_customer is None and not self.request.user.has_perm(APP_NAME+".add_cartitem") :
            message="دسترسی غیر مجاز"
            return result,message,cart_item,cart_items
         
        if me_customer is None:
            message=" 22دسترسی غیر مجاز"
            return result,message,cart_item,cart_items

        if 'shop_id' in kwargs:
            shop_id=kwargs["shop_id"]
         
        if 'quantity' in kwargs:
            quantity=kwargs["quantity"] 

            if quantity is None:
                quantity=1

        cart_item=CartItem.objects.filter(shop_id=shop_id).filter(customer_id=me_customer.id).first()
        if cart_item is None:
            cart_item=CartItem(customer_id=me_customer.id,quantity=0) 

        cart_item.shop_id=shop_id
        cart_item.quantity=quantity
        (result,message,cart_item)=cart_item.save() 
        cart_items=CartItem.objects.filter(customer_id=me_customer.id)

        return result,message,cart_item,cart_items
 
 
   
    def change_cart_item(self,*args,**kwargs):
        result,message,cart_item,cart_items=FAILED,"",None,[]
        me_customer=CustomerRepo(request=self.request).me
        if  me_customer is None and not self.request.user.has_perm(APP_NAME+".change_cartitem") :
            message="دسترسی غیر مجاز"
            return result,message,cart_item,cart_items
         
        if me_customer is None:
            message=" 22دسترسی غیر مجاز"
            return result,message,cart_item,cart_items
        cart_item=CartItem(customer_id=me_customer.id) 
        if 'shop_id' in kwargs:
            shop_id=kwargs["shop_id"]
         
        if 'quantity' in kwargs:
            quantity=kwargs["quantity"] 

        cart_item=CartItem.objects.filter(shop_id=shop_id).filter(customer_id=me_customer.id).first()
        if cart_item is None:
            cart_item=CartItem(shop_id=shop_id,customer_id=me_customer.id,quantity=0)
        cart_item.quantity=quantity

        (result,message,cart_item)=cart_item.save() 
        cart_items=CartItem.objects.filter(customer_id=me_customer.id)
        return result,message,cart_item,cart_items
 
 
    def checkout(self,customer_id,*args,**kwargs):
        result,message,invoices=FAILED,"",[]
         

        if self.request.user.has_perm(APP_NAME+".add_cartitem"):
                pass
                
        else :
            me_customer=CustomerRepo(request=self.request).me
            if me_customer is None:
                
                message="دسترسی غیر مجاز"
                return result,message,invoices
            elif customer_id==me_customer.id:
                pass
            else:
                message="دسترسی غیر مجاز"
                return result,message,invoices
        invoices=[]
        cart_items=kwargs['cart_items']
        suppliers_ids=[]
        for cart_item in cart_items:
            shop=Shop.objects.filter(pk=cart_item['shop_id']).first()
            if shop is not None:
                supplier_id=shop.supplier.id
                if supplier_id not in suppliers_ids:
                    suppliers_ids.append(supplier_id)
        customer=CustomerRepo(request=self.request).customer(pk=customer_id)
        if customer is None:
            return FAILED,"مشتری نادرست انتخاب شده است",[]
        for supplier_id in suppliers_ids:

            supplier=Supplier.objects.filter(pk=supplier_id).first()
            from accounting.repo import InvoiceRepo,Invoice,InvoiceLine
            from django.utils import timezone
            invoice_data={}
            invoice_data['bedehkar_id']=customer.person_account.id
            invoice_data['bestankar_id']=supplier.person_account.id
            invoice_data['title']="فاکتور جدید"
            invoice_data['amount']=0
            if 'address' in kwargs:
                invoice_data['address']=kwargs['address']
            if 'postal_code' in kwargs:
                invoice_data['postal_code']=kwargs['postal_code']
            invoice_data['event_datetime']=timezone.now()

            invoice=Invoice(**invoice_data)
            invoice.save()

            log_data={}
            from log.repo import LogRepo
            log_data['person_id']=PersonRepo(request=self.request).me.id
            log_data['url']=invoice.get_absolute_url()
            log_data['title']="ذخیره فاکتور جدید"
            log_data['description']="فاکتور جدید"
            log_data['app_name']=APP_NAME
            LogRepo(request=self.request).add_log(**log_data)

            # invoice.save()
            invoices.append(invoice)
            for cart_item in cart_items:
                shop=Shop.objects.filter(pk=cart_item['shop_id']).first()
                if shop.supplier.id==supplier_id:
                    if cart_item['quantity']>shop.available:
                        message="تعداد درخواستی از موجودی بیشتر است."
                        return FAILED,message,None
                    invoice_line=InvoiceLine()
                    invoice_line.discount_percentage=shop.discount_percentage
                    invoice_line.invoice_id=invoice.id
                    invoice_line.invoice_line_item_id=shop.product_id
                    invoice_line.quantity=cart_item['quantity'] 
                    invoice_line.unit_price=shop.unit_price
                    invoice_line.unit_name=shop.unit_name
                    invoice_line.save() 
                    shop.available-=cart_item['quantity']
                    shop.save()
                    CartItem.objects.filter(shop_id=shop.id).filter(customer_id=customer.id).delete()
        result=SUCCEED
        links=''
        for invoice in invoices:
            links+=f""" <br><a target="_blank" href="{invoice.get_print_url()}">{invoice.title}</a> """
        message=f"""{len(invoices)} فاکتور برای شما با موفقیت ایجاد شد"""+links
        return result,message,invoices
 

    def checkout_old_temp_must_be_deleted(self,*args,**kwargs):
        result,message,invoices=FAILED,"",[]
        me_customer=CustomerRepo(request=self.request).me
        if  me_customer is None and not self.request.user.has_perm(APP_NAME+".add_cartitem") :
            message="دسترسی غیر مجاز"
            return result,message,invoices
         
        if me_customer is None:
            message=" 22دسترسی غیر مجاز"
            return result,message,invoices
        customer_id=me_customer.id
        invoices=[]
        cart_items=CartItem.objects.filter(customer_id=customer_id)
        suppliers_ids=[]
        for cart_item in cart_items:
            supplier_id=cart_item.shop.supplier.id
            if supplier_id not in suppliers_ids:
                suppliers_ids.append(supplier_id)
        customer=CustomerRepo(request=self.request).customer(pk=customer_id)
        if customer is None:
            return FAILED,"مشتری نادرست انتخاب شده است",[]
        for supplier_id in suppliers_ids:

            supplier=SupplierRepo(request=self.request).supplier(pk=supplier_id)    
            from accounting.repo import InvoiceRepo,Invoice,InvoiceLine
            from django.utils import timezone
            invoice_data={}
            invoice_data['bedehkar_id']=customer.person_account.id
            invoice_data['bestankar_id']=supplier.person_account.id
            invoice_data['title']="فاکتور جدید"
            invoice_data['amount']=0
            invoice_data['event_datetime']=timezone.now()
            # invoice=Invoice(invoice_data)
            invoice=Invoice(**invoice_data)
            invoice.save()

            # invoice.save()
            invoices.append(invoice)
            for cart_item in cart_items:
                if cart_item.shop.supplier.id==supplier_id:
                    invoice_line=InvoiceLine()
                    invoice_line.invoice_id=invoice.id
                    invoice_line.invoice_line_item_id=cart_item.shop.product_id
                    invoice_line.quantity=cart_item.quantity
                    invoice_line.unit_price=cart_item.shop.unit_price
                    invoice_line.unit_name=cart_item.shop.unit_name
                    invoice_line.save() 
                    cart_item.delete()
        result=SUCCEED
        links=''
        for invoice in invoices:
            links+=f""" <br><a target="_blank" href="{invoice.get_print_url()}">{invoice.title}</a> """
        message=f"""{len(invoices)} فاکتور برای شما با موفقیت ایجاد شد"""+links
        return result,message,invoices
 

class SupplierRepo():
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        self.objects=Supplier.objects.filter(pk=0)
        me_person=PersonRepo(request=request).me
        me_shipper=ShipperRepo(request=request).me
        me_supplier=Supplier.objects.filter(person_account__person_id=me_person.id).first()
        self.me=me_supplier
        if me_shipper is not None:
            self.objects=me_shipper.suppliers.all()
        if me_supplier is not None:
            self.objects=Supplier.objects.filter(id=me_supplier.id) 
        if request.user.has_perm(APP_NAME+".view_supplier"):
            self.objects=Supplier.objects.all()

    def list(self,*args, **kwargs):
        objects=self.objects
         
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]

            objects=objects.filter(Q(name__contains=search_for) | Q(code=search_for) | Q(pure_code=pure_code ) )
        return objects.all()
    

    def supplier(self,*args, **kwargs):
        if "supplier_id" in kwargs and kwargs["supplier_id"] is not None:
            return self.objects.filter(pk=kwargs['supplier_id']).first() 
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "code" in kwargs and kwargs["code"] is not None:
            return self.objects.filter(code=kwargs['code']).first()
             
        if "barcode" in kwargs and kwargs["barcode"] is not None:
            a= self.objects.filter(barcode=kwargs['barcode']).first() 
            return a 
           


 
    def add_supplier(self,*args,**kwargs):
        result,message,supplier=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_supplier") :
            message="دسترسی غیر مجاز"
            return result,message,supplier
        

        supplier=Supplier() 
        if 'person_account_id' in kwargs and kwargs['person_account_id'] is not None and kwargs['person_account_id']>0:
            person_account_id=kwargs["person_account_id"]
            supplier.person_account_id=person_account_id
            if len(Supplier.objects.filter(person_account_id=person_account_id))>0:
                message='برای این حساب قبلا فروشنده ایجاد شده است.'
                return result,message,None
         
        (result,message,supplier)=supplier.save() 

        if 'region_id' in kwargs:
            supplier.regions.add(kwargs["region_id"])
        return result,message,supplier
 

class CustomerGroupRepo():
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        self.objects=CustomerGroup.objects
 

    def list(self,*args, **kwargs):
        objects=self.objects
        pure_code="876454453342236"
        try:
            pure_code=int(kwargs["search_for"]) 
        except:
            pass
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]

            objects=objects.filter(Q(name__contains=search_for) | Q(code=search_for) | Q(pure_code=pure_code ) )
        return objects.all()
    

    def customer_group(self,*args, **kwargs):
        if "customer_group_id" in kwargs and kwargs["customer_group_id"] is not None:
            return self.objects.filter(pk=kwargs['customer_group_id']).first() 
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "code" in kwargs and kwargs["code"] is not None:
            return self.objects.filter(code=kwargs['code']).first()
             
        if "barcode" in kwargs and kwargs["barcode"] is not None:
            a= self.objects.filter(barcode=kwargs['barcode']).first() 
            return a 
           


 
    def add_customer_group(self,*args,**kwargs):
        result,message,customer_group=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_customer_group") :
            message="دسترسی غیر مجاز"
            return result,message,customer_group
        

        customer_group=CustomerGroup() 
        if 'name' in kwargs:
            customer_group.name=kwargs["name"]
            if len(CustomerGroup.objects.filter(name=kwargs["name"]))>0:
                message='این گروه قبلا ایجاد شده است.'
                return result,message,None
 
        
        (result,message,customer_group)=customer_group.save() 

        return result,message,customer_group
 