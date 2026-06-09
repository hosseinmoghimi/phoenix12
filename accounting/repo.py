from .models import Asset,Category,FinancialDocument,FinancialDocumentLine,InvoiceLineItemUnit,Cheque
from .models import InvoiceLine,InvoiceLineItem,Account,Product,Service,FinancialEvent,FinancialYear
from .models import Invoice,Bank,PersonCategory,FinancialYear,PersonAccount,ProductSpecification
from .models import BankAccount
from .models import Brand
from .apps import APP_NAME
from .enums import *
from log.repo import LogRepo
from django.db.models import Q
from django.shortcuts import reverse
from authentication.repo import PersonRepo
# from processmanagement.permission import Permission,OperationEnum
from utility.num import filter_number
from utility.calendar import PersianCalendar
from utility.constants import FAILED,SUCCEED
from .constants import EXCEL_PRODUCTS_DATA_START_ROW,EXCEL_SERVICES_DATA_START_ROW,EXCEL_ACCOUNTS_DATA_START_ROW
from .defaults import default_accounts,default_persons,default_banks
from .enums import AccountTypeEnum,AccountNatureEnum
from .server_settings import ACCOUNT_LEVEL_NAMES
from authentication.models import Person

from utility.repo import leolog,Repo
 
class InvoiceLineItemUnitRepo:
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        
        
        self.objects=None
        if request.user.has_perm(APP_NAME+'.view_invoicelineitemunit'):
            self.objects=InvoiceLineItemUnit.objects
        else:
            self.objects=InvoiceLineItemUnit.objects.filter(pk=0)

    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            objects=objects.filter(title__contains=kwargs['search_for']) 
            
        if "invoice_line_item_id" in kwargs:
            objects=objects.filter(invoice_line_item_id=kwargs['invoice_line_item_id']) 
        return objects.all()
    def product_unit(self,*args, **kwargs):
        if "product_unit_id" in kwargs:
            return self.objects.filter(pk=kwargs['product_unit_id']).first() 
        if "pk" in kwargs:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs:
            return self.objects.filter(pk=kwargs['id']).first() 

                     
    def add_invoice_line_item_unit(self,*args, **kwargs):
        invoice_line_item_unit,message,result=(None,"",FAILED)
        # if not Permission(request=self.request).is_permitted(APP_NAME,OperationEnum.ADD,"productprice"):
        if not self.request.user.has_perm(APP_NAME+".add_invoicelineitemunit"):
            message="دسترسی غیر مجاز"
            return result,message,invoice_line_item_unit
        
 
        invoice_line_item_unit=InvoiceLineItemUnit()
        if 'invoice_line_item_id' in kwargs:
            invoice_line_item_unit.invoice_line_item_id=kwargs['invoice_line_item_id']

        if 'service' in kwargs:
            invoice_line_item_unit.invoice_line_item_id=kwargs['service']
        
        if 'product_id' in kwargs:
            invoice_line_item_unit.invoice_line_item_id=kwargs['product_id']
        if 'unit_price' in kwargs:
            invoice_line_item_unit.unit_price=kwargs['unit_price']
        if 'unit_name' in kwargs:
            invoice_line_item_unit.unit_name=kwargs['unit_name']
        if 'default' in kwargs:
            invoice_line_item_unit.default=kwargs['default']
        if 'coef' in kwargs:
            invoice_line_item_unit.coef=kwargs['coef']
        if invoice_line_item_unit.unit_price<1:
            message='قیمت را صفر انتخاب کرده اید.'
            return result,message,invoice_line_item_unit
        invoice_line_item_unit.save()
        result=SUCCEED
        message="قیمت جدید با موفقیت اضافه گردید."
         
        invoice_line_item_units=InvoiceLineItemUnit.objects.filter(invoice_line_item_id=invoice_line_item_unit.invoice_line_item.id)
        return result,message,invoice_line_item_units


class InvoiceLineRepo:
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me_person=None
    
        from authentication.repo import PersonRepo
        me_person=PersonRepo(request=request).me
        self.me_person=me_person
        
        self.objects=None
        if request.user.has_perm(APP_NAME+".view_invoiceline"):
            self.objects=InvoiceLine.objects
        elif request.user.is_authenticated:
            accs=[]
            if me_person is not None:
                my_accounts=AccountRepo(request=request).my_accounts
                for acc in my_accounts:
                    accs.append(acc.id)
            self.objects=InvoiceLine.objects.filter(person_id=me_person.id)
        else:
            self.objects=InvoiceLine.objects.filter(pk=0)

    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            objects=objects.filter(title__contains=kwargs['search_for']) 
        if "invoice_line_item_id" in kwargs:
            objects=objects.filter(invoice_line_item_id=kwargs['invoice_line_item_id']) 
        return objects.all()
    def invoice_line(self,*args, **kwargs):
        if "invoice_line_id" in kwargs:
            return self.objects.filter(pk=kwargs['invoice_line_id']).first() 
        if "pk" in kwargs:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs:
            return self.objects.filter(pk=kwargs['id']).first() 

            
        
    def add_invoice_line(self,*args,**kwargs):
        result,message,meal=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_invoiceline"):
            message="دسترسی غیر مجاز"
            return result,message,meal

        invoice_line=InvoiceLine(person_id=self.me_person.id)
        if 'invoice_line_item_id' in kwargs:
            invoice_line_item_id=kwargs["invoice_line_item_id"]
            invoice_line.invoice_line_item_id=invoice_line_item_id

        if 'invoice_id' in kwargs:
            invoice_line.invoice_id=kwargs["invoice_id"]
            invoice=invoice_line.invoice
            if invoice.status==FinancialEventStatusEnum.APPROVED:
                message='فاکتور تایید شده و امکان تغییر ، ویرایش و افزودن سطر وجود ندارد.'
                return FAILED,message,None
            
            if invoice.status==FinancialEventStatusEnum.DELIVERED:
                message='فاکتور تحویل شده و امکان تغییر ، ویرایش و افزودن سطر وجود ندارد.'
                return FAILED,message,None
            
            if invoice.status==FinancialEventStatusEnum.FINISHED:
                message='فاکتور نهایی شده و امکان تغییر ، ویرایش و افزودن سطر وجود ندارد.'
                return FAILED,message,None
            
        if 'description' in kwargs:
            invoice_line.description=kwargs["description"]
        if 'status' in kwargs:
            invoice_line.status=kwargs["status"]
        if 'discount_percentage' in kwargs:
            invoice_line.discount_percentage=kwargs["discount_percentage"]
        if 'quantity' in kwargs:
            invoice_line.quantity=kwargs["quantity"]
        if 'unit_price' in kwargs:
            unit_price=kwargs["unit_price"]
            invoice_line.unit_price=unit_price

        if 'unit_name' in kwargs:
            unit_name=kwargs["unit_name"]
            invoice_line.unit_name=unit_name
        if 'save' in kwargs or kwargs["default_price"]:
            save=kwargs["save"]
            if save or kwargs["default_price"]:
                if 'coef' in kwargs:
                    coef=kwargs["coef"]
                if 'default_price' in kwargs:
                    default_price=kwargs["default_price"]
                InvoiceLineItemUnitRepo(request=self.request).add_invoice_line_item_unit(
                    invoice_line_item_id=invoice_line_item_id,
                    coef=coef,
                    default=default_price,
                    unit_name=unit_name,
                    unit_price=unit_price,
                    )
        invoice_line.row=len(invoice_line.invoice.invoiceline_set.all())+1
        result,message,invoice_line=invoice_line.save()
 
        return result,message,invoice_line
           

class AccountRepo(Repo):
    def __init__(self,request,*args, **kwargs):
        super(AccountRepo,self).__init__(request=request,app_name=APP_NAME,*args, **kwargs)
        self.request=request
        self.objects=Account.objects.filter(id=0)
        me_person=self.person 
        if request.user.has_perm(APP_NAME+".view_account"):
            self.objects=Account.objects.all()
        elif me_person is not None:
            self.objects=self.my_accounts
    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            codeee=str(filter_number(search_for))
            objects=objects.filter(Q(title__contains=search_for) | Q(code=search_for) | Q(code=codeee) )
        if "parent_id" in kwargs:
            parent_id=kwargs["parent_id"]
            objects=objects.filter(parent_id=parent_id)  
        if "level" in kwargs:
            level=kwargs["level"]
            objects=objects.filter(level=level)  
        return objects.all()
    @property
    def my_accounts(self,*args, **kwargs):
        return PersonAccount.objects.filter(person_id=self.me_person.id)

    def roots(self,*args, **kwargs):
        objects=self.objects.filter(parent_id=None)
        return objects.all()
    
    def merge_account(self,*args, **kwargs):
           
        result,message,merged_account=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+'.change_account'):
            message='شما مجوز دسترسی برای این عملکرد را ندارید.'
            return FAILED,message,None
        deleting_account=self.account(pk=kwargs['deleting_account_id'])
        updating_account=self.account(pk=kwargs['updating_account_id'])

        i=0
        from core.models import Page
        from attachments.models import Comment,Image,Link,Download,Location
        for child in Page.objects.filter(parent_id=deleting_account.id):
            i+=1
            child.parent_id=updating_account.id
            child.save()
        message+='<br>'+f'( {i} )'+'فرزند با موفقیت همگام سازی شد.'    
 
        i=0
        for comment in Comment.objects.filter(page_id=deleting_account.id):
            i+=1
            comment.page_id=updating_account.id
            comment.save()
        message+='<br>'+f'( {i} )'+'کامنت ها با موفقیت همگام سازی شد.'    
        i=0
        for link in Link.objects.filter(page_id=deleting_account.id):
            i+=1
            link.page_id=updating_account.id
            link.save()
        message+='<br>'+f'( {i} )'+'لینک ها با موفقیت همگام سازی شد.'    
            
        i=0
        for download in Download.objects.filter(page_id=deleting_account.id):
            i+=1
            download.page_id=updating_account.id
            download.save()
        message+='<br>'+f'( {i} )'+'دانلود ها با موفقیت همگام سازی شد.'    



        for image in Image.objects.filter(page_id=deleting_account.id):
            i+=1
            image.page_id=updating_account.id
            image.save()
        message+='<br>'+f'( {i} )'+'تصاویر با موفقیت همگام سازی شد.'    



        i=0
        for financial_document_line in FinancialDocumentLine.objects.filter(account_id=deleting_account.id):
            i+=1
            financial_document_line.account_id=updating_account.id
            financial_document_line.save()
        message+='<br>'+f'( {i} )'+'سطر های اسناد مالی با موفقیت همگام سازی شد.'    

            
        from warehouse.models import WareHouse
        i=0
        for warehouse in WareHouse.objects.filter(person_account_id=deleting_account.id):
            i+=1
            warehouse.person_account_id=updating_account.id
            warehouse.save()
        if i>0:
            message+='<br>'+f'( {i} )'+'انبارها با موفقیت همگام سازی شد.'    
        

        from transport.models import ServiceMan
        i=0
        for service_man in ServiceMan.objects.filter(person_account_id=deleting_account.id):
            i+=1
            service_man.person_account_id=updating_account.id
            service_man.save()
        if i>0:
            message+='<br>'+f'( {i} )'+'تعمیرکار ها با موفقیت همگام سازی شد.'    
        


        from organization.models import OrganizationalUnit
        i=0
        for organizational_unit in OrganizationalUnit.objects.filter(person_account_id=deleting_account.id):
            i+=1
            organizational_unit.person_account_id=updating_account.id
            organizational_unit.save()
        if i>0:
            message+='<br>'+f'( {i} )'+'واحد های سازمانی با موفقیت همگام سازی شد.'    
        
        from market.models import MarketPerson
        i=0
        for market_person in MarketPerson.objects.filter(person_account_id=deleting_account.id):
            i+=1
            market_person.person_account_id=updating_account.id
            market_person.save()
        if i>0:
            message+='<br>'+f'( {i} )'+'افراد فروشگاه با موفقیت همگام سازی شد.'    
        

            
        from health.models import Doctor,Patient
        i=0
        for doctor in Doctor.objects.filter(person_account_id=deleting_account.id):
            i+=1
            doctor.person_account_id=updating_account.id
            doctor.save()
        if i>0:
            message+='<br>'+f'( {i} )'+'دکتر ها با موفقیت همگام سازی شد.'    
        

        i=0
        for patient in Patient.objects.filter(person_account_id=deleting_account.id):
            i+=1
            patient.person_account_id=updating_account.id
            patient.save()
        if i>0:
            message+='<br>'+f'( {i} )'+'بیمار ها با موفقیت همگام سازی شد.'    
        

        from school.models import School,Teacher,Student
        i=0
        for school in School.objects.filter(person_account_id=deleting_account.id):
            i+=1
            school.person_account_id=updating_account.id
            school.save()
        if i>0:
            message+='<br>'+f'( {i} )'+'مدرسه ها با موفقیت همگام سازی شد.'    
        


        i=0
        for teacher in Teacher.objects.filter(person_account_id=deleting_account.id):
            i+=1
            teacher.person_account_id=updating_account.id
            teacher.save()
        if i>0:
            message+='<br>'+f'( {i} )'+'دبیر ها با موفقیت همگام سازی شد.'    
        


        i=0
        for student in Student.objects.filter(person_account_id=deleting_account.id):
            i+=1
            student.person_account_id=updating_account.id
            student.save()
        if i>0:
            message+='<br>'+f'( {i} )'+'دانش آموزان با موفقیت همگام سازی شد.'    
        


        i=0
        for financial_event in FinancialEvent.objects.filter(Q(bedehkar_id=deleting_account.id)|Q(bestankar_id=deleting_account.id)):
            i+=1
            if financial_event.bedehkar_id==deleting_account.id:
                financial_event.bedehkar_id=updating_account.id
            if financial_event.bestankar_id==deleting_account.id:
                financial_event.bestankar_id=updating_account.id
            financial_event.save()
        message+='<br>'+f'( {i} )'+'رویدادهای مالی با موفقیت همگام سازی شد.'    

        result=SUCCEED
        message+='<br>'+'با موفقیت همگام سازی شد.'    
        return result,message,merged_account
 
    def normalize_all_accounts(self,*args, **kwargs):
        result,message,counter=FAILED,'',0
        if not self.request.user.has_perm(APP_NAME+".change_account"):
            message="دسترسی غیر مجاز"
            message='شما مجوز دسترسی به این عملکرد را ندارید.'
            return result,message,counter
        
        # for account in Account.objects.all():
        #     account.bedehkar=0
        #     account.bestankar=0
        #     account.balance=0
        #     account.save()
        

        sw=1
        if sw==1:
            for account in Account.objects.filter(parent_id=None):
                result,message,counter2=account.normalize_to_top()
                counter+=counter2

        if sw==2:
            for account in Account.objects.all():
                counter+=1
                account.normalize()
            result=SUCCEED
            message=f'{counter} حساب مالی با موفقیت نرمال سازی شد.'
        
        return result,message,counter
        
    def import_accounts_from_excel(self,*args,**kwargs):
        result,message,accounts=FAILED,"",[]
        excel_file=kwargs['excel_file']
        # import pandas
        
        # df = pandas.read_excel(excel_file)
        # accounts=[]
        # for row in df.columns[0]:
        #     print (df.columns)
        import openpyxl 

        wb = openpyxl.load_workbook(excel_file)
        try:
            ws = wb['accounts']
        
        except:
            message='فایل شما برگه حساب ها ندارد.'
            return result,message,None
        count=kwargs['count']
        try:
            count=int(ws.cell(row=1, column=2).value)
        except:
            
            message='فایل برگه حساب ها ، تعداد ندارد.'
            return result,message,None  
        accounts_to_import=[]
        START_ROW=EXCEL_ACCOUNTS_DATA_START_ROW
        for i in range(START_ROW,count+START_ROW):
            account={}
            
            
            i=str(i) 
            # product['id']=ws['A'+str(i)].value
            iiiddd=ws['C'+i].value
            if iiiddd is not None:
                id=int(ws['C'+i].value)
                parent_code=(ws['B'+i].value)
                if parent_code is not None:
                    parent_code=(parent_code)
                code=(ws['D'+i].value)
                title=(ws['E'+i].value)
                color=(ws['F'+i].value)
                thumbnail_origin=(ws['G'+i].value)

                account['id']=id
                account['parent_code']=parent_code
                account['code']=code
                account['title']=title
                account['color']=color
                account['thumbnail_origin']=thumbnail_origin
                if account['title'] is not None and not account['title']=="":
                    accounts_to_import.append(account) 
        modified=added=0
        for account in accounts_to_import:
            old_account=Account.objects.filter(title=account["title"]).filter(code=account["code"]).first()
            if old_account is not None:
                old_account.title=account["title"]
                # old_account.unit_name=account["unit_name"]
                old_account.thumbnail_origin=account["thumbnail_origin"]
                # old_account.unit_price=account["unit_price"] 
                # old_account.thumbnail_origin=account["thumbnail_origin"] 
                old_account.save()
                modified+=1
            else:
                try:
                    result,message,new_account=self.add_account(title=account["title"],
                                                                code=account["code"],
                                                                parent_code=account["parent_code"],
                                                                id=account["id"],
                                                                color=account["color"],
                                                                thumbnail_origin=account["thumbnail_origin"] ,
                                                                )
                    accounts.append(new_account)
                except:
                    pass
                # new_account.title=account["title"]
                # new_account.barcode=account["barcode"]
                # new_account.unit_name=account["unit_name"]
                # new_account.unit_price=account["unit_price"] 
                # new_account.save()
                if result==SUCCEED:
                    added+=1
        result=SUCCEED
        message=f"""{added} حساب اضافه شد.
                    <br>
                    {modified} حساب ویرایش شد. """
        accounts=self.list()


        if True:
            log_data={}
            from log.repo import LogRepo
            log_data['person_id']=PersonRepo(request=self.request).me.id
            log_data['url']=reverse("accounting:accounts")
            log_data['title']="بازیابی حساب ها"
            log_data['description']=message
            log_data['app_name']=APP_NAME
            LogRepo(request=self.request).add_log(**log_data)
        
        return result,message,accounts



    def account(self,*args, **kwargs):
        if "account_id" in kwargs and kwargs["account_id"] is not None:
            return self.objects.filter(pk=kwargs['account_id']).first()  
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "code" in kwargs and kwargs["code"] is not None:
            code=kwargs['code']
            code=kwargs['code']

            account= self.objects.filter(code=code).first()
            return account
        if "account_code" in kwargs and kwargs["account_code"] is not None:
            a= self.objects.filter(code=kwargs['account_code']).first() 
            if a is not None:
                return a
            else:
                try:
                    a= self.objects.filter(pure_code=filter_number(kwargs['account_code'])).first() 
                    if a is not None:
                        return a
                except:
                    pass
       

           
    def set_account_parent(self,*args, **kwargs):
        result,message,account,parent=FAILED,"",None,None
        account=self.account(*args,**kwargs)
        if account is None:
            message='حساب پیدا نشد.'
            return result,message,None,None
        parent=None
        parent_code=kwargs["parent_code"]
        if parent_code=='0':
            account.parent=None
        else:    
            parent=self.account(account_code=parent_code)
            account.parent_id=parent.id
        account.save()
        result=SUCCEED
        message="با موفقیت تغییر یافت"
        return result,message,account,parent


    def initial_default_accounts(self,*args, **kwargs):
        account_group_counter=0
        basic_accounts_counter=0
        moein_accounts_counter=0
        moein2_accounts_counter=0
        tafsili_accounts_counter=0 

        result=SUCCEED
        message=""  
        if not self.request.user.has_perm(APP_NAME+".add_account"):
            message="دسترسی غیر مجاز"
            result=FAILED
            return message,result
        
        accounts=default_accounts() 

        for account in accounts:
            parent_account=None
            if 'parent_code' in account:
                parent_account=Account.objects.filter(code=account["parent_code"]).first()
            new_account=Account(title=account["name"],color=account["color"],code=account['code'],priority=account['priority'],parent=parent_account)
            # new_account=Account(parent=parent_account,**kwargs)
            new_account.save()
            # account_group_counter+=1
            # basic_accounts_counter+=1   
 
                                  
        accounts=Account.objects
        account_group_counter=len(accounts.filter(type=AccountTypeEnum.GROUP))
        basic_accounts_counter=len(accounts.filter(type=AccountTypeEnum.BASIC))
        moein_accounts_counter=len(accounts.filter(type=AccountTypeEnum.MOEIN_1))
        moein2_accounts_counter=len(accounts.filter(type=AccountTypeEnum.MOEIN_2))
        tafsili_accounts_counter=len(accounts.filter(Q(type=AccountTypeEnum.TAFSILI_1)|Q(type=AccountTypeEnum.TAFSILI_2)|Q(type=AccountTypeEnum.TAFSILI_3)|Q(type=AccountTypeEnum.TAFSILI_4)))
                                                        

        if result==SUCCEED:
            message="با موفقیت اضافه گردید."
        message+=f"<br>{account_group_counter}   گروه حساب" 
        message+=f"<br>{basic_accounts_counter}   حساب  کل " 
        message+=f"<br>{moein_accounts_counter}  حساب معین سطح یک " 
        message+=f"<br>{moein2_accounts_counter}  حساب معین سطح دو " 
        message+=f"<br>{tafsili_accounts_counter}  حساب تفصیلی " 

        me_person=PersonRepo(request=self.request).me
        new_log={}
        new_log['title']="افزودن حساب های پیش فرض"
        new_log['app_name']=APP_NAME
        new_log['person']=me_person
        new_log['url']=reverse("accounting:accounts")
        new_log['description']="حساب های پیش فرض اضافه شدند."
        LogRepo(request=self.request).add_log(**new_log)
        return result,message
 
    def delete_all_accounts(self,*args, **kwargs):
        
        result,message=FAILED,''
        if not self.request.user.has_perm(APP_NAME+".delete_account"):
            message="دسترسی غیر مجاز"
            return message,result
        FinancialDocumentLine.objects.all().delete()
        FinancialEvent.objects.all().delete()
        FinancialDocument.objects.all().delete()
        # TafsiliAccount.objects.all().delete()
        # MoeinAccount.objects.all().delete()
        # BasicAccount.objects.all().delete()
        # AccountGroup.objects.all().delete() 
        Account.objects.all().delete() 
        result=SUCCEED
        message="همه حساب ها حذف شد."

        me_person=PersonRepo(request=self.request).me
        new_log={}
        new_log['title']="حذف همه حساب ها"
        new_log['app_name']=APP_NAME
        new_log['person']=me_person
        new_log['description']="همه ی حساب ها حذف شدند."
        LogRepo(request=self.request).add_log(**new_log)

        return result,message
    
    def add_account(self,*args,**kwargs):
        result,message,account=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_account"):
            message="دسترسی غیر مجاز"
            return result,message,account

        account=Account()
        if 'title' in kwargs:
            account.title=kwargs["title"]
        if 'parent_id' in kwargs:
            if kwargs["parent_id"]>0:
                account.parent_id=kwargs["parent_id"]
        if 'color' in kwargs:
            account.color=kwargs["color"]
        if 'code' in kwargs:
            account.code=kwargs["code"]
        if 'priority' in kwargs:
            account.priority=kwargs["priority"]
        if 'type' in kwargs:
            account.type=kwargs["type"]

            
        if 'parent_code' in kwargs:
            parent_code= kwargs["parent_code"]
            parent=Account.objects.filter(code=parent_code).first()
            if parent is not None:
                account.parent_id=parent.id

        if 'nature' in kwargs:
            account.nature=kwargs["nature"]
        (result,message,account)=account.save()
        new_log={}
        new_log['title']='حساب مالی جدید'
        new_log['url']=account.get_absolute_url()
        new_log['description']=message
        self.log(**new_log)
        return result,message,account
 

class PersonAccountRepo(Repo):
    def __init__(self,request,*args, **kwargs):
        super(PersonAccountRepo,self).__init__(request,app_name=APP_NAME,*args, **kwargs)
        self.objects=PersonAccount.objects.filter(pk=0)
        me_person=self.me
        if request.user.has_perm(APP_NAME+'.view_personaccount'):
            self.objects=PersonAccount.objects.all().order_by('person__full_name')
        elif me_person is not None:
            self.objects=PersonAccount.objects.filter(person__user_id=me_person.user.id).order_by('person__full_name')
    def list(self,*args, **kwargs):
        objects=self.objects
        pure_code="876454453342236"
        try:
            pure_code=int(kwargs["search_for"]) 
        except:
            pass
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]

            objects=objects.filter(Q(category__contains=search_for))

        if "category" in kwargs:
            category=kwargs["category"]
            objects=objects.filter(Q(category=category)   )
        if "person_id" in kwargs:
            person_id=kwargs["person_id"]
            objects=objects.filter(Q(person_id=person_id)   )
        return objects.all()
  
    def add_person_account(self,*args, **kwargs):
        result,message,person_account=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_personaccount"):
            message="دسترسی غیر مجاز"
            return result,message,person_account
        person_account=PersonAccount()
        
        
    

        if 'person' in kwargs:
            person_account.person=kwargs['person']

        if 'nature' in kwargs:
            person_account.nature=kwargs['nature']
        if 'person_id' in kwargs:
            person_account.person_id=kwargs['person_id']
        if 'person_category' in kwargs:
            person_account.person_category=kwargs['person_category']
        if 'person_category_id' in kwargs:
            person_account.person_category_id=kwargs['person_category_id']
        (result,message,person_account)=person_account.save()
        if person_account is None:
            return result,message,None
        else:
            new_log={}
            new_log['title']='حساب مالی شخص جدید'
            new_log['url']=person_account.get_absolute_url()
            new_log['description']=message
            self.log(**new_log)
        # result=SUCCEED
        # message="با موفقیت حساب فرد ایجاد شد."
        # if 'title' in kwargs:
        #     person_account.title=kwargs['title']
        # if 'code' in kwargs:
        #     code=kwargs['code']
        #     person_account.code=code
        if 'color' in kwargs:
            person_account.color=kwargs['color']
        
            person_account.save()
            
        return result,message,person_account

    def person_account(self,*args, **kwargs):
        if "person_account_id" in kwargs and kwargs["person_account_id"] is not None:
            return self.objects.filter(pk=kwargs['person_account_id']).first() 
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "code" in kwargs and kwargs["code"] is not None:
            code=kwargs['code']
            return self.objects.filter(code=code).first()
        if "person_account_code" in kwargs and kwargs["person_account_code"] is not None:
            return self.objects.filter(code=kwargs['person_account_code']).first()
             
        if "account_code" in kwargs and kwargs["account_code"] is not None:
            a= self.objects.filter(code=kwargs['account_code']).first() 
            if a is not None:
                return a
            else:
                try:
                    a= self.objects.filter(pure_code=filter_number(kwargs['account_code'])).first() 
                    if a is not None:
                        return a
                except:
                    pass
                    
    def delete_all(self,*args,**kwargs):
        
        result,message=FAILED,''
        if not self.request.user.has_perm(APP_NAME+".delete_personaccount"):
            message="دسترسی غیر مجاز"
            return result,message
        PersonAccount.objects.all().delete() 
                   
        result=SUCCEED
        message="همه حذف شدند."
        return result,message
     

    def add_account_to_person(self,*args,**kwargs):
        result,message,account,person,action=FAILED,"",None,None,None
        if not self.request.user.has_perm(APP_NAME+".add_personaccount"):
            message="دسترسی غیر مجاز"
            return result,message,account,person,action
        person_account=PersonAccount()
        
        person_account.person_id=kwargs['person_id']
        person_account.person_category_id=kwargs['person_category_id']

        result,message,person_account=person_account.save()
        if result==FAILED:
            person_account=None
            return result,message,person_account,None,"ALREADY_EXISTED"

        action="ADDED"
        result=SUCCEED
        message=f"حساب مالی  {person_account.name} با موفقیت به شخص {person_account.person.full_name} اضافه شد."
            
        return result,message,person_account,person_account.person,action

    def remove_account_from_person(self,*args,**kwargs):
        result,message,person_account_id=FAILED,"",0
        if not self.request.user.has_perm(APP_NAME+".add_personaccount"):
            message="دسترسی غیر مجاز"
            return result,message,person_account_id
             
        person_id=kwargs['person_id']
        person_category_id=kwargs['person_category_id']

        person_account=PersonAccount.objects.filter(person_id=person_id).filter(person_category_id=person_category_id).first()
        if person_account is None:
            result=FAILED
            message="حساب مالی وجود ندارد."
            return result,message,person_account_id

        person_account_id=person_account.id
        try:
            person_account.delete()
            person_account_=PersonAccount.objects.filter(id=person_account_id).first()
            if person_account_ is None:
                result=SUCCEED
                message=f"حساب مالی  {person_account.name} با موفقیت از شخص {person_account.person.full_name} حذف شد."
        except:
            message="حذف نشد. "+"ابتدا رویداد های مالی مرتبط را حذف کنید."+"تا تراز فرد صفر شده و هیچ سندی با شخص در ارتباط نباشد."
            return result,message,person_account_id

        return result,message,person_account_id

 
class FinancialYearRepo():
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        self.objects=FinancialYear.objects

        
        self.objects=None
        if request.user.has_perm(APP_NAME+".view_financialyear"):
            self.objects=FinancialYear.objects
        elif request.user.is_authenticated:

            self.objects=FinancialYear.objects.filter(id=0)
                 
        else:
            self.objects=FinancialYear.objects.filter(pk=0)
 
    def list(self,*args, **kwargs):
        objects=self.objects
        pure_code="876454453342236"
        try:
            pure_code=int(kwargs["search_for"]) 
        except:
            pass
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]

            objects=objects.filter(Q(name__contains=search_for) | Q(code=search_for) | Q(pure_code=pure_code ) )
        return objects.all()
     
    def financial_year(self,*args, **kwargs):
        if "financial_year_id" in kwargs and kwargs["financial_year_id"] is not None:
            return self.objects.filter(pk=kwargs['financial_year_id']).first()
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "date" in kwargs and kwargs["date"] is not None:
            return self.objects.filter(start_date_lte=kwargs['date']).filter(end_date_gte=kwargs['date']).first() 
        
    def current_financial_year(self):
        return self.objects.filter(in_progress=True).first()
   
    def add_financial_year(self,*args,**kwargs):
        result,message,financial_year,financial_years=FAILED,"",None,[]
        if not self.request.user.has_perm(APP_NAME+".add_financialyear"):
            message="دسترسی غیر مجاز"
            return result,message,financial_year,financial_years

        financial_year=FinancialYear() 

 

    
        if 'start_date' in kwargs:
            year=kwargs['start_date'][:2]
            if year=="13" or year=="14":
                kwargs['start_date']=PersianCalendar().to_gregorian(kwargs["start_date"])
            financial_year.start_date=kwargs['start_date']

            
        if 'end_date' in kwargs:
            year=kwargs['end_date'][:2]
            if year=="13" or year=="14":
                kwargs['end_date']=PersianCalendar().to_gregorian(kwargs["end_date"])
            financial_year.end_date=kwargs['end_date']

 

        if 'name' in kwargs:
            financial_year.name=kwargs["name"] 

        if 'description' in kwargs:
            financial_year.description=kwargs["description"] 

        if 'status' in kwargs:
            financial_year.status=kwargs["status"] 
 
        if(len(FinancialYear.objects.filter(name=financial_year.name))>0):
            financial_year=None
            message="نام وارد شده تکراری است."
            result=FAILED
            return result,message,financial_year,financial_years

 
        (result,message,financial_year)=financial_year.save()
        if result==FAILED:
            return result,message,financial_year,financial_years
        
        financial_years=FinancialYear.objects.order_by("start_date")
 
        return result,message,financial_year,financial_years
 
      
class PersonCategoryRepo():
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        self.objects=PersonCategory.objects
    def list(self,*args, **kwargs):
        objects=self.objects
 
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]

            objects=objects.filter(Q(name__contains=search_for) | Q(code=search_for) | Q(pure_code=pure_code ) )
        return objects.all()
     
    def edit_person_category(self,*args, **kwargs):
        result,message,person_category=FAILED,'',None
        if not self.request.user.has_perm(APP_NAME+".change_personcategory"):
            message="دسترسی غیر مجاز"
            return result,message,None

        person_category=self.person_category(person_category_id=kwargs['person_category_id'])
        account=AccountRepo(request=self.request).account(account_id=kwargs['account_id'])
        if account is None or person_category is None:
            message="داده های مرتبط پیدا نشد."
            return result,message,None
        person_category.account_id=account.id
        person_category.title=kwargs['title']
        person_category.code_length=kwargs['code_length']
        person_category.save()
        message='با موفقیت اصلاح شد.'
        return SUCCEED,message,person_category

    def delete_all(self,*args,**kwargs):
        result,message=FAILED,''
        if not self.request.user.has_perm(APP_NAME+".delete_personcategory"):
            message="دسترسی غیر مجاز"
            return result,message
        PersonCategory.objects.all().delete()
        
        result=SUCCEED
        message="همه حذف شدند."
        return result,message

    def person_category(self,*args, **kwargs):
        if "person_category_id" in kwargs and kwargs["person_category_id"] is not None:
            return self.objects.filter(pk=kwargs['person_category_id']).first()
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "code" in kwargs and kwargs["code"] is not None:
            return self.objects.filter(code=kwargs['code']).first()
        if "title" in kwargs and kwargs["title"] is not None:
            return self.objects.filter(title=kwargs['title']).first()
             
        if "account_code" in kwargs and kwargs["account_code"] is not None:
            a= self.objects.filter(code=kwargs['account_code']).first() 
            if a is not None:
                return a
            else:
                try:
                    a= self.objects.filter(pure_code=filter_number(kwargs['account_code'])).first() 
                    if a is not None:
                        return a
                except:
                    pass
       
    def add_person_category(self,*args,**kwargs):
        result,message,person_category=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_personcategory"):
            message="دسترسی غیر مجاز"
            return result,message,person_category
         
        person_category=PersonCategory()
          
        if 'code_length' in kwargs:
            code_length=kwargs["code_length"]
            person_category.code_length=code_length
            
        if 'account_id' in kwargs:
            account_id=kwargs["account_id"]
            person_category.account=Account.objects.filter(id=account_id).first()

          
        if 'account_code' in kwargs:
            account_code=kwargs["account_code"]
            person_category.account=Account.objects.filter(code=account_code).first()

          
         
        if 'title' in kwargs:
            if len(PersonCategory.objects.filter(title=kwargs["title"]))>0:
                message="نام وارد شده تکراری می باشد."
                return result,message,None
            person_category.title=kwargs["title"]
        if 'priority' in kwargs:
            person_category.priority=kwargs["priority"] 
        if 'code_length' in kwargs:
            person_category.code_length=kwargs["code_length"] 
        person_category.save()
        result=SUCCEED
        message="دسته بندی جدید برای اشخاص با موفقیت اضافه گردید."
        # (result,message,person_category)=person_category.save()
        # (result,message,person_category)=person_category.save()
        return result,message,person_category

    def initial_default_person_categories(self,*args,**kwargs):
        result=SUCCEED
        message=""
        person_categories,persons=default_persons()
        for person_category in person_categories:
            result,message,new_person_category=self.add_person_category(**person_category)
                
        for person in persons:
            from authentication.repo import PersonRepo
            result,message,new_person=PersonRepo(request=self.request).add_person(**person)

        return result,message       


class ProductSpecificationRepo:
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        
        
        self.objects=None
        if request.user.has_perm(APP_NAME+".view_event"):
            self.objects=ProductSpecification.objects
        elif request.user.is_authenticated:
            accs=[]
            for person in Person.objects.filter(user_id=request.user.id):
                my_accounts=AccountRepo(request=request).my_accounts
                for acc in my_accounts:
                    accs.append(acc.id)
            self.objects=Event.objects.filter(Q(bedehkar_id__in=accs)|Q(bestankar_id__in=accs))
        else:
            self.objects=Event.objects.filter(pk=0)

    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            objects=objects.filter(title__contains=kwargs['search_for']) 
        return objects.all()
    def product_specification(self,*args, **kwargs):
        if "product_specification_id" in kwargs:
            return self.objects.filter(pk=kwargs['product_specification_id']).first() 
        if "pk" in kwargs:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs:
            return self.objects.filter(pk=kwargs['id']).first() 

                     
    def add_product_specification(self,*args, **kwargs):
        product_specification,message,result,deleted_id=(None,"",FAILED,0)
        # if not Permission(request=self.request).is_permitted(APP_NAME,OperationEnum.ADD,"productspecification"):
        if not self.request.user.has_perm(APP_NAME+".add_account"):
            message="دسترسی غیر مجاز"
            return result,message,product_specification,deleted_id
        # if len(Account.objects.filter(title=kwargs['title']))>0:
        #     message="از قبل حسابی با همین عنوان ثبت شده است."
        #     return product_specification,message,result
        product_specification=ProductSpecification()
        
        if 'product_id' in kwargs:
            product_specification.product_id=kwargs['product_id']
        if 'name' in kwargs:
            product_specification.name=kwargs['name']
        if 'value' in kwargs:
            product_specification.value=kwargs['value'] 
        deleted=ProductSpecification.objects.filter(product_id=product_specification.product_id).filter(name=product_specification.name).first()
        if deleted is not None:
            deleted_id=deleted.id
            deleted.delete()
         
        product_specification.save()
        result=SUCCEED
        message="ویژگی جدید با موفقیت اضافه گردید."
         
 
        return result,message,product_specification,deleted_id


class ProductRepo():
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        self.objects=Product.objects
    def list(self,*args, **kwargs):
        objects=self.objects
        
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]

            objects=objects.filter(Q(title__contains=search_for) | Q(barcode=search_for)|Q(model__contains=search_for))
        if "for_home" in kwargs:
            for_home=kwargs["for_home"]
            if for_home:
                objects=objects.filter(priority=0)
                
        if "title" in kwargs:
            title=kwargs["title"]

            objects=objects.filter(Q(title__contains=title) | Q(barcode=title)|Q(model__contains=title))
            
        if "id__in" in kwargs:
            id__in=kwargs["id__in"]

            objects=objects.filter(id__in=id__in) 

        return objects.all()
    
    def delete_all(self):
        if self.request.user.has_perm(APP_NAME+".delete_product"):
            Product.objects.all().delete()
            result=SUCCEED
            message="همه کالاها با موفقیت حذف شدند."
        return result,message
    def add_product_to_category(self,*args, **kwargs):
        
        result,message,category,product_categories=FAILED,"",None,[]
        if not self.request.user.has_perm(APP_NAME+".change_product"):
            message="دسترسی غیر مجاز"
            return result,message,category,product_categories
        

        product=Product.objects.filter(pk=kwargs['product_id']).first()
        category=Category.objects.filter(pk=kwargs['category_id']).first()
        if category is not None and product is not None:
            if product in category.products.all():
                category.products.remove(product.id)
                message="حذف شد"
            else:
                category.products.add(product.id)
                message="اضافه شد"
            result=SUCCEED
            product_categories=product.category_set.all()
        return result,message,category,product_categories
    
    def merge_product(self,*args, **kwargs):
           
        result,message,merged_product=FAILED,"",None
        deleting_product=self.product(pk=kwargs['deleting_product_id'])
        updating_product=self.product(pk=kwargs['updating_product_id'])

        i=0
        for invoice_line in InvoiceLine.objects.filter(invoice_line_item_id=deleting_product.id):
            i+=1
            invoice_line.invoice_line_item_id=updating_product.id
            invoice_line.save()
        message+='<br>'+f'( {i} )'+'سطر فاکتور ها با موفقیت همگام سازی شد.'    

            
        i=0
        for invoice_line_item_unit in InvoiceLineItemUnit.objects.filter(invoice_line_item_id=deleting_product.id):
            i+=1
            invoice_line_item_unit.invoice_line_item_id=updating_product.id
            invoice_line_item_unit.save()
        message+='<br>'+f'( {i} )'+'قیمت ها با موفقیت همگام سازی شد.'    
        
        
        i=0
        for product_specification in ProductSpecification.objects.filter(product_id=deleting_product.id):
            i+=1
            product_specification.product_id=updating_product.id
            product_specification.save()
        message+='<br>'+f'( {i} )'+'ویژگی ها با موفقیت همگام سازی شد.'    

        i=0

        
        from projectmanager.models import RemoteClient
        
        i=0
        for remote_client in RemoteClient.objects.filter(product_id=deleting_product.id):
            i+=1
            remote_client.product_id=updating_product.id
            remote_client.save()
        if i>0:
            message+='<br>'+f'( {i} )'+'ریموت کلاینت ها با موفقیت همگام سازی شد.'    

        from market.models import Shop
        
        i=0
        for shop in Shop.objects.filter(product_id=deleting_product.id):
            i+=1
            shop.product_id=updating_product.id
            shop.save()
        if i>0:
            message+='<br>'+f'( {i} )'+'ریموت کلاینت ها با موفقیت همگام سازی شد.'    



        from attachments.models import Image,Link,Download,Comment,Location
         
        i=0
        for image in Image.objects.filter(page_id=deleting_product.id):
            i+=1
            image.page_id=updating_product.id
            image.save()
        message+='<br>'+f'( {i} )'+'تصویر ها با موفقیت همگام سازی شد.'    

        i=0
        for link in Link.objects.filter(page_id=deleting_product.id):
            i+=1
            link.page_id=updating_product.id
            link.save()
        message+='<br>'+f'( {i} )'+'لینک ها با موفقیت همگام سازی شد.'    

        i=0
        for download in Download.objects.filter(page_id=deleting_product.id):
            i+=1
            download.page_id=updating_product.id
            download.save()
        message+='<br>'+f'( {i} )'+'دانلود ها با موفقیت همگام سازی شد.'    

        i=0
        for comment in Comment.objects.filter(page_id=deleting_product.id):
            i+=1
            comment.page_id=updating_product.id
            comment.save()
        message+='<br>'+f'( {i} )'+'کامنت ها با موفقیت همگام سازی شد.'    
 
        # for location in Location.objects.filter(page_id=deleting_product.id):
        #     location.page_id=updating_product.id
        #     message+='<br>'+'موقعیت ها با موفقیت همگام سازی شد.'    
        #     location.save()

        result=SUCCEED
        message+='<br>'+'با موفقیت همگام سازی شد.'    
        return result,message,merged_product


        return result,message,merged_product 
    
    def product(self,*args, **kwargs):
        product=None
        if "product_id" in kwargs and kwargs["product_id"] is not None:
            product= self.objects.filter(pk=kwargs['product_id']).first()
            if product is not None:
                return product 
        if "barcode" in kwargs and kwargs["barcode"] is not None:
            a= self.objects.filter(barcode=kwargs['barcode']).first() 
            if product is None and a is not None:
                product= a 
                return a
        
        if "pk" in kwargs and kwargs["pk"] is not None:
            if product is None:
                product= self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            if product is None:
                product= self.objects.filter(pk=kwargs['id']).first() 
        if "code" in kwargs and kwargs["code"] is not None:
            if product is None:
                product= self.objects.filter(barcode=kwargs['code']).first()
             
        return product
       
    def add_product(self,*args,**kwargs):
        result,message,product=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_product"):
            message="دسترسی غیر مجاز"
            return result,message,product
        if len(Product.objects.filter(title=kwargs["title"]))>0:
            message="نام تکراری برای کالای جدید"
            return result,message,product

        product=Product() 
        if 'title' in kwargs:
            product.title=kwargs["title"]

        if 'id' in kwargs:
            product.id=kwargs["id"]

            
        if 'brand_id' in kwargs:
            brand_id=kwargs['brand_id']
            if brand_id is not None and brand_id>0:
                product.brand_id=brand_id
        
            
        if 'model' in kwargs and kwargs['model'] is not None:
            product.model=kwargs["model"]
            
        if 'thumbnail_origin' in kwargs and kwargs['thumbnail_origin'] is not None:
            product.thumbnail_origin=kwargs["thumbnail_origin"]
         
        if 'barcode' in kwargs and kwargs["barcode"] is not None and not kwargs["barcode"]=="":
            product.barcode=kwargs["barcode"]
        
        if product.barcode is not None and len(product.barcode)>0:
            
            if len(Product.objects.filter(barcode=product.barcode))>0:
                message="بارکد تکراری برای کالای جدید"
                return result,message,None

        if 'rop' in kwargs  and kwargs['rop'] is not None:
            rop=kwargs['rop']
            product.rop=rop
        (result,message,product)=product.save()
        if 'unit_price' in kwargs and kwargs['unit_price'] is not None:
            if 'unit_name' in kwargs and kwargs['unit_name'] is not None:
                if 'coef' in kwargs and kwargs['coef'] is not None:
                    ili_unit=InvoiceLineItemUnit()
                    ili_unit.unit_name=kwargs["unit_name"]
                    ili_unit.coef=kwargs["coef"]
                    ili_unit.unit_price=kwargs["unit_price"]
                    ili_unit.invoice_line_item_id=product.id
                    ili_unit.default=True
                    ili_unit.save()
        else:
            pass
            ili_unit=InvoiceLineItemUnit()
            ili_unit.unit_name=UnitNameEnum.ADAD
            ili_unit.coef=1
            ili_unit.unit_price=0
            ili_unit.invoice_line_item_id=product.id
            ili_unit.default=True
            ili_unit.save()
                 


        if 'category_id' in kwargs:
            category_id=kwargs['category_id']
            if category_id is not None and category_id>0:
                category=Category.objects.filter(pk=category_id).first()
                if category is not None:
                    category.products.add(product.id)
        return result,message,product
 

    def import_products_from_excel(self,*args,**kwargs):
        result,message,products=FAILED,"",[]
        excel_file=kwargs['excel_file']
        # import pandas
        
        # df = pandas.read_excel(excel_file)
        # products=[]
        # for row in df.columns[0]:
        #     print (df.columns)
        import openpyxl 

        wb = openpyxl.load_workbook(excel_file)
        try:
            ws = wb['products']
        except:
            message='فایل شما برگه محصولات ندارد.'
            return result,message,None
        count=kwargs['count']
        try:
            count=int(ws.cell(row=1, column=2).value)
        except:
            message='فایل برگه محصولات ، تعداد ندارد.'
            return result,message,None 

        products_to_import=[]
        START_ROW=EXCEL_PRODUCTS_DATA_START_ROW

        for i in range(START_ROW,count+START_ROW):
            product={}
            i=str(i) 
            # product['id']=ws['A'+str(i)].value
            iiiddd=ws['B'+i].value
            if iiiddd is not None:
                id=int(ws['B'+i].value)
                title=(ws['C'+i].value)
                barcode=(ws['D'+i].value)
                unit_name=(ws['E'+i].value)
                unit_price=int(ws['F'+i].value)
                thumbnail_origin=(ws['G'+i].value)
                category_id=(ws['H'+i].value)
                product['id']=id
                product['title']=title
                product['barcode']=barcode
                product['unit_name']=unit_name
                product['unit_price']=unit_price
                product['thumbnail_origin']=thumbnail_origin
                product['category_id']=category_id
                # product['thumbnail_origin']=ws['F'+str(i)].value
                if product['title'] is not None and not product['title']=="":
                    products_to_import.append(product) 
        modified=added=0
        for product in products_to_import:
            category=None
            if product['category_id'] is not None:
                category=Category.objects.filter(pk=product['category_id']).first()
            old_product=Product.objects.filter(title=product["title"]).filter(barcode=product["barcode"]).first()
            if old_product is not None:
                old_product.title=product["title"]
                # old_product.unit_name=product["unit_name"]
                old_product.thumbnail_origin=product["thumbnail_origin"]
                # old_product.unit_price=product["unit_price"] 
                # old_product.thumbnail_origin=product["thumbnail_origin"] 
                old_product.save()
                
                if category is not None:
                    category.products.add(old_product.id)

                modified+=1
            else:
                try:
                    result,message,new_product=self.add_product(title=product["title"],barcode=product["barcode"],unit_name=product["unit_name"],unit_price=product["unit_price"],thumbnail_origin=product["thumbnail_origin"] ,coef=1)
                    products.append(new_product)
                    
                    if category is not None:
                        category.products.add(new_product.id)

                except:
                    pass
                # new_product.title=product["title"]
                # new_product.barcode=product["barcode"]
                # new_product.unit_name=product["unit_name"]
                # new_product.unit_price=product["unit_price"] 
                # new_product.save()
                if result==SUCCEED:
                    added+=1
        result=SUCCEED
        message=f"""{added} محصول اضافه شد.
                    <br>
                    {modified} محصول ویرایش شد. """
        products=self.list()

        
        
        if True:
            log_data={}
            from log.repo import LogRepo
            log_data['person_id']=PersonRepo(request=self.request).me.id
            log_data['url']=reverse("accounting:products")
            log_data['title']="بازیابی کالا ها"
            log_data['description']=message
            log_data['app_name']=APP_NAME
            LogRepo(request=self.request).add_log(**log_data)

        return result,message,products


class InvoiceLineItemRepo():
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        self.objects=InvoiceLineItem.objects
    def list(self,*args, **kwargs):
        objects=self.objects
        
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]

            objects=objects.filter(Q(title__contains=search_for))
        return objects.all()
    
         
    def invoice_line_item(self,*args, **kwargs):
        if "invoice_line_item_id" in kwargs and kwargs["invoice_line_item_id"] is not None:
            return self.objects.filter(pk=kwargs['invoice_line_item_id']).first() 
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "code" in kwargs and kwargs["code"] is not None:
            return self.objects.filter(barcode=kwargs['code']).first()
             
        if "barcode" in kwargs and kwargs["barcode"] is not None:
            a= self.objects.filter(barcode=kwargs['barcode']).first() 
            return a 
           
    def add_invoice_line_item(self,*args,**kwargs):
        result,message,invoice_line_item=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_invoice_line_item"):
            message="دسترسی غیر مجاز"
            return result,message,invoice_line_item
        if len(InvoiceLineItem.objects.filter(title=kwargs["title"]))>0:
            message="نام تکراری برای مورد جدید"
            return result,message,invoice_line_item

        invoice_line_item=InvoiceLineItem() 

        if 'title' in kwargs:
            invoice_line_item.title=kwargs["title"]
        if 'unit_price' in kwargs:
            invoice_line_item.unit_price=kwargs["unit_price"]
        if 'unit_price' in kwargs:
            invoice_line_item.unit_price=kwargs["unit_price"]
            
        if 'barcode' in kwargs and kwargs["barcode"] is not None and not kwargs["barcode"]=="":
            invoice_line_item.barcode=kwargs["barcode"]
        
        if 'unit_name' in kwargs:
            invoice_line_item.unit_name=kwargs["unit_name"]

            
        if invoice_line_item.barcode is not None and len(invoice_line_item.barcode)>0:
            
            if len(Product.objects.filter(barcode=invoice_line_item.barcode))>0:
                message="بارکد تکراری برای کالای جدید"
                return result,message,None

        (result,message,invoice_line_item)=invoice_line_item.save()
        if 'category_id' in kwargs:
            category_id=kwargs["category_id"]
            category=Category.objects.filter(pk=category_id).first()
            if category is not None:
                category.invoice_line_items.add(invoice_line_item.id)
        coef=1
        if 'coef' in kwargs:
            coef=kwargs["coef"]
        if invoice_line_item.unit_price>0:
            InvoiceLineItemUnitRepo(request=self.request).add_invoice_line_item_unit(invoice_line_item_id=invoice_line_item.id,unit_price=invoice_line_item.unit_price,unit_name=invoice_line_item.unit_name,coef=coef)
        return result,message,invoice_line_item
 

class BankAccountRepo():
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        me_person=PersonRepo(request=request).me
        if request.user.has_perm(APP_NAME+'.view_bankaccount'):
            self.objects=BankAccount.objects.all()
        else:
            self.objects=BankAccount.objects.filter(id=0)
    def list(self,*args, **kwargs):
        objects=self.objects
  
        if "bank_id" in kwargs:
            bank_id=kwargs["bank_id"]
            objects=objects.filter(Q(bank_id=bank_id))
  
        if "person_id" in kwargs:
            person_id=kwargs["person_id"]
            objects=objects.filter(Q(person_id=person_id))
       
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            objects=objects.filter(Q(title__contains=search_for) | Q(account_no__contains=search_for) | Q(card_no__contains=search_for) | Q(shaba_no__contains=search_for)  )
        return objects.all()
     
    def bank_account(self,*args, **kwargs):
        if "bank_account_id" in kwargs and kwargs["bank_account_id"] is not None:
            return self.objects.filter(pk=kwargs['bank_account_id']).first()
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "code" in kwargs and kwargs["code"] is not None:
            return self.objects.filter(code=kwargs['code']).first()
    
    
    def add_bank_account(self,*args, **kwargs):
        result,message,bank_account=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_bankaccount"):
            message="دسترسی غیر مجاز"
            return result,message,bank_account
        bank_account=BankAccount()
        
        if 'person' in kwargs:
            bank_account.person=kwargs['person']
        if 'person_id' in kwargs:
            bank_account.person_id=kwargs['person_id']
        if 'person_category' in kwargs:
            bank_account.person_category=kwargs['person_category']
        if 'person_category_id' in kwargs:
            bank_account.person_category_id=kwargs['person_category_id']
        # result=SUCCEED
        # message="با موفقیت حساب فرد ایجاد شد."
        if 'title' in kwargs:
            bank_account.title=kwargs['title']
        if 'code' in kwargs:
            code=kwargs['code']
            bank_account.code=code
        if 'color' in kwargs:
            bank_account.color=kwargs['color']
        
  
        if 'card_no' in kwargs:
            bank_account.card_no=kwargs["card_no"] 

            
  
        if 'shaba_no' in kwargs:
            bank_account.shaba_no=kwargs["shaba_no"] 

            
  
        if 'account_no' in kwargs:
            bank_account.account_no=kwargs["account_no"] 
  
        if 'title' in kwargs:
            bank_account.title=kwargs["title"] 

        if 'bank_id' in kwargs:
            bank_account.bank_id=kwargs["bank_id"] 
        # result,message,bank_account=bank_account.save()
        # return result,message,bank_account 
        return bank_account.save()
 
  
class BrandRepo():
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        self.objects=Brand.objects
       

    def list(self,*args, **kwargs):
        objects=self.objects
  
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            objects=objects.filter(Q(name__contains=search_for) )
        return objects.all()
     
    def brand(self,*args, **kwargs):
        if "brand_id" in kwargs and kwargs["brand_id"] is not None:
            return self.objects.filter(pk=kwargs['brand_id']).first()
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "code" in kwargs and kwargs["code"] is not None:
            return self.objects.filter(code=kwargs['code']).first()
    
    def initial_default_brands(self,*args, **kwargs):
      
        brands_counter=0 

        result=SUCCEED
        message=""  
        if not self.request.user.has_perm(APP_NAME+".add_brand"):
            message="دسترسی غیر مجاز"
            result=FAILED
            return message,result
        for brand in default_brands():
            new_brand=Brand(name=brand["name"])
            # new_account=Account(parent=parent_account,**kwargs)
            new_brand.save()
            brands_counter+=1

        message=f"{brands_counter} بانک با موفقیت اضافه شد."
        return result,message

        
    def delete_all(self,*args, **kwargs):
      
        brands_counter=0 

        result=SUCCEED
        message=""  
        if not self.request.user.has_perm(APP_NAME+".delete_brand"):
            message="دسترسی غیر مجاز"
            result=FAILED
            return message,result
        for brand in Brand.objects.all():
            brand.delete()
            brands_counter+=1

        message=f"<p>{brands_counter} بانک با موفقیت حذف شد.</p>"
        return result,message
    
    
    def add_brand(self,*args,**kwargs):
        result,message,brand=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_brand"):
            message="دسترسی غیر مجاز"
            return result,message,brand

        brand=Brand() 

 

        if Brand.objects.filter(name=kwargs['name']).first() is not None:
            message="نام وارد شده تکراری است."
            brand=None
            return result,message,brand

   
        if 'name' in kwargs:
            brand.name=kwargs["name"] 
        brand.save()       
        message="برند جدید با موفقیت اضافه شد."
        result=SUCCEED
        return result,message,brand


class AssetRepo():
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        self.objects=Asset.objects
       

    def list(self,*args, **kwargs):
        objects=self.objects
  
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            objects=objects.filter(Q(full_name__contains=search_for) | Q(melli_code__contains=search_for) | Q(code=search_for))
        return objects.all()
     
    def asset(self,*args, **kwargs):
        if "asset_id" in kwargs and kwargs["asset_id"] is not None:
            return self.objects.filter(pk=kwargs['asset_id']).first()
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "code" in kwargs and kwargs["code"] is not None:
            return self.objects.filter(code=kwargs['code']).first()
     
        
    def delete_all(self,*args, **kwargs):
      
        assets_counter=0 

        result=SUCCEED
        message=""  
        if not self.request.user.has_perm(APP_NAME+".delete_asset"):
            message="دسترسی غیر مجاز"
            result=FAILED
            return message,result
        for asset in Asset.objects.all():
            asset.delete()
            assets_counter+=1

        message=f"<p>{assets_counter} بانک با موفقیت حذف شد.</p>"
        return result,message
    
    
    def add_asset(self,*args,**kwargs):
        result,message,asset=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_asset"):
            message="دسترسی غیر مجاز"
            return result,message,asset

        asset=Asset() 

 

        if Asset.objects.filter(title=kwargs['title']).first() is not None:
            message="نام وارد شده تکراری است."
            asset=None
            return result,message,asset

  
        if 'title' in kwargs:
            asset.title=kwargs["title"] 

        if 'owner_id' in kwargs:
            asset.owner_id=kwargs["owner_id"]
            
        asset.save()       
        message="دارایی جدید با موفقیت اضافه شد."
        result=SUCCEED
        return result,message,asset
 

class BankRepo():
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        self.objects=Bank.objects
       
        me_person=PersonRepo(request=request).me
        if request.user.has_perm(APP_NAME+'.view_bank'):
            self.objects=Bank.objects.all()
        else:
            self.objects=Bank.objects.filter(id=0)

    def list(self,*args, **kwargs):
        objects=self.objects
  
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            objects=objects.filter(Q(full_name__contains=search_for) | Q(melli_code__contains=search_for) | Q(code=search_for))
        return objects.all()
     
    def bank(self,*args, **kwargs):
        if "bank_id" in kwargs and kwargs["bank_id"] is not None:
            return self.objects.filter(pk=kwargs['bank_id']).first()
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "code" in kwargs and kwargs["code"] is not None:
            return self.objects.filter(code=kwargs['code']).first()
    
    def initial_default_banks(self,*args, **kwargs):
      
        banks_counter=0 

        result=SUCCEED
        message=""  
        if not self.request.user.has_perm(APP_NAME+".add_bank"):
            message="دسترسی غیر مجاز"
            result=FAILED
            return message,result
        for bank in default_banks():
            new_bank=Bank(name=bank["name"])
            # new_account=Account(parent=parent_account,**kwargs)
            new_bank.save()
            banks_counter+=1

        message=f"{banks_counter} بانک با موفقیت اضافه شد."
        return result,message

        
    def delete_all(self,*args, **kwargs):
      
        banks_counter=0 

        result=SUCCEED
        message=""  
        if not self.request.user.has_perm(APP_NAME+".delete_bank"):
            message="دسترسی غیر مجاز"
            result=FAILED
            return message,result
        for bank in Bank.objects.all():
            bank.delete()
            banks_counter+=1

        message=f"<p>{banks_counter} بانک با موفقیت حذف شد.</p>"
        return result,message
    
    
    def add_bank(self,*args,**kwargs):
        result,message,bank=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_bank"):
            message="دسترسی غیر مجاز"
            return result,message,bank

        bank=Bank() 

 

        if Bank.objects.filter(name=kwargs['name']).first() is not None:
            message="نام وارد شده تکراری است."
            bank=None
            return result,message,bank

  
        if 'name' in kwargs:
            bank.name=kwargs["name"] 
        bank.save()       
        message="بانک جدید با موفقیت اضافه شد."
        result=SUCCEED
        return result,message,bank

 
class ServiceRepo():
    def __init__(self,request,*args, **kwargs):
        self.request=request
        self.me=None
        self.objects=Service.objects
    def list(self,*args, **kwargs):
        objects=self.objects
        
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]

            objects=objects.filter(Q(title__contains=search_for))
        return objects.all()
    
    def add_product_to_category(self,*args, **kwargs):
        
        result,message,category,product_categories=FAILED,"",None,[]
        if not self.request.user.has_perm(APP_NAME+".change_product"):
            message="دسترسی غیر مجاز"
            return result,message,category,product_categories
        

        product=Product.objects.filter(pk=kwargs['product_id']).first()
        category=Category.objects.filter(pk=kwargs['category_id']).first()
        if category is not None and product is not None:
            if product in category.products.all():
                category.products.remove(product.id)
                message="حذف شد"
            else:
                category.products.add(product.id)
                message="اضافه شد"
            result=SUCCEED
            product_categories=product.category_set.all()
        return result,message,category,product_categories
       
    def service(self,*args, **kwargs):
        if "service_id" in kwargs and kwargs["service_id"] is not None:
            return self.objects.filter(pk=kwargs['service_id']).first() 
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "code" in kwargs and kwargs["code"] is not None:
            return self.objects.filter(barcode=kwargs['code']).first()
             
        if "barcode" in kwargs and kwargs["barcode"] is not None:
            a= self.objects.filter(barcode=kwargs['barcode']).first() 
            return a 
            
    def add_service(self,*args,**kwargs):
        result,message,service=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_service"):
            message="دسترسی غیر مجاز"
            return result,message,service
        if len(Service.objects.filter(title=kwargs["title"]))>0:
            message="نام تکراری برای سرویس جدید"
            return result,message,service

        service=Service() 

        if 'title' in kwargs:
            service.title=kwargs["title"]
         
    

        (result,message,service)=service.save()
        if 'unit_price' in kwargs:
            if 'unit_name' in kwargs:
                if 'coef' in kwargs:
                    ili_unit=InvoiceLineItemUnit()
                    ili_unit.unit_name=kwargs["unit_name"]
                    ili_unit.coef=kwargs["coef"]
                    ili_unit.unit_price=kwargs["unit_price"]
                    ili_unit.invoice_line_item_id=service.id
                    ili_unit.default=True
                    ili_unit.save()

                 

        if 'category_id' in kwargs:
            pass
            # category_id=kwargs["category_id"]
            # category=Category.objects.filter(pk=category_id).first()
            # if category is not None:
            #     category.services.add(service.id)
        coef=1 
        return result,message,service
 
    def import_services_from_excel(self,*args,**kwargs):
        result,message,services=FAILED,"",[]
        excel_file=kwargs['excel_file']
        # import pandas
        
        # df = pandas.read_excel(excel_file)
        # services=[]
        # for row in df.columns[0]:
        #     print (df.columns)
        import openpyxl 

        wb = openpyxl.load_workbook(excel_file)
        try:
            ws = wb['services']
        except:
            message='فایل شما برگه خدمات ندارد.'
            return result,message,None
        count=kwargs['count']
        try:
            count=int(ws.cell(row=1, column=2).value)
        except:
            
            message='فایل برگه خدمات ، تعداد ندارد.'
            return result,message,None 

        services_to_import=[]

        START_ROW=EXCEL_SERVICES_DATA_START_ROW
        for i in range(START_ROW,count+START_ROW):
            service={}
            
            i=str(i) 
            # product['id']=ws['A'+str(i)].value
            iiiddd=ws['B'+i].value
            if iiiddd is not None:
                id=int(ws['B'+i].value)
                title=(ws['C'+i].value)
                unit_name=(ws['D'+i].value)
                unit_price=int(ws['E'+i].value)
                thumbnail_origin=(ws['F'+i].value)
                service['id']=id
                service['title']=title
                service['unit_name']=unit_name
                service['unit_price']=unit_price
                service['thumbnail_origin']=thumbnail_origin
                # service['thumbnail_origin']=ws['F'+str(i)].value
                if service['title'] is not None and not service['title']=="":
                    services_to_import.append(service) 
        modified=added=0
        for service in services_to_import:
            old_service=Service.objects.filter(title=service["title"]).first()
            if old_service is not None:
                old_service.title=service["title"]
                # old_service.unit_name=service["unit_name"]
                old_service.thumbnail_origin=service["thumbnail_origin"]
                # old_service.unit_price=service["unit_price"] 
                # old_service.thumbnail_origin=service["thumbnail_origin"] 
                old_service.save()
                modified+=1
            else:
                result,message,new_service=self.add_service(title=service["title"],unit_name=service["unit_name"],unit_price=service["unit_price"] ,coef=1)
                # new_service.title=service["title"]
                # new_service.barcode=service["barcode"]
                # new_service.unit_name=service["unit_name"]
                # new_service.unit_price=service["unit_price"] 
                # new_service.save()
                if result==SUCCEED:
                    added+=1
        result=SUCCEED
        message=f"""{added} سرویس اضافه شد.
                    <br>
                    {modified} سرویس ویرایش شد. """
        services=self.list()

        
        if True:
            log_data={}
            from log.repo import LogRepo
            log_data['person_id']=PersonRepo(request=self.request).me.id
            log_data['url']=reverse("accounting:services")
            log_data['title']="بازیابی سرویس ها"
            log_data['description']=message
            log_data['app_name']=APP_NAME
            LogRepo(request=self.request).add_log(**log_data)

        return result,message,services


class FinancialDocumentRepo(Repo):
    def __init__(self,request,*args, **kwargs):
        super(FinancialDocumentRepo,self).__init__(request=request,app_name=APP_NAME,*args, **kwargs)
        self.objects=[]

        if self.me is not None:
            if request.user.has_perm('.view_financialdocument'):
                self.objects=FinancialDocument.objects
            else:
                me_accounts=PersonAccount.objects.filter(person_id=self.me.id)
                self.objects=FinancialDocument.objects.filter(pk=0)
    def list(self,*args, **kwargs):
        objects=self.objects
        if "financial_year_id" in kwargs:
            objects=objects.filter(financial_year_id=kwargs['financial_year_id']) 
        if "search_for" in kwargs:
            objects=objects.filter(title__contains=kwargs['search_for']) 
        return objects.all()
    
    def normalize_all_financial_documents(self,*args, **kwargs):
        result,message,counter=FAILED,'',0
        if not self.request.user.has_perm(APP_NAME+".change_financialdocument"):
            message="دسترسی غیر مجاز"
            message='شما مجوز دسترسی به این عملکرد را نداریـــد.'
            return result,message,counter
        for financial_document in FinancialDocument.objects.all():
            result,message=financial_document.normalize()
            counter+=1
        message=f'{counter} سند مالی با موفقیت نرمال سازی شد.'
        return result,message,counter
    
    
    def financial_document(self,*args, **kwargs):
        if "financial_document_id" in kwargs:
            return self.objects.filter(pk=kwargs['financial_document_id']).first() 
        if "pk" in kwargs:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs:
            return self.objects.filter(pk=kwargs['id']).first() 

            
    def add_financial_document(self,*args, **kwargs):
        result,message,financial_document=FAILED,"",None
        # if not Permission(request=self.request).is_permitted(APP_NAME,OperationEnum.ADD,"financialdocument"):
        if not self.request.user.has_perm(APP_NAME+".add_financialdocument"):
            message="دسترسی غیر مجاز"
            return result,message,financial_document
        

        old=FinancialDocument.objects.filter(title=kwargs['title']).first()
        if old is not None:
            message='از قبل سندی با همین عنوان موجود می باشد و نمی توان با نام تکراری سندی ایجاد کرد.'
            return FAILED,message,None
        f_year=FinancialYear.objects.filter(status=FinancialYearStatusEnum.IN_PROGRESS).first()
        if f_year is None:
            url=reverse(APP_NAME+":financial_years")
            message="سال مالی فعال وجود ندارد. ابتدا ایجاد کنید."+"<br>"+"<a href='"+url+"'>سال های مالی</a>"
            return result,message,financial_document

        financial_document=FinancialDocument(financial_year_id=f_year.id)

        if 'title' in kwargs:
            financial_document.title=kwargs['title']
     
        
        # if 'financial_year_id' in kwargs:
        #     payment.financial_year_id=kwargs['financial_year_id']
        # else:
        #     payment.financial_year_id=FinancialYear.get_by_date(date=payment.transaction_datetime).id
        result,message,financial_document=financial_document.save()
        if result==SUCCEED:
            message="با موفقیت اضافه شد."
            me_person=PersonRepo(request=self.request).me
            new_log={}
            new_log['title']="سند مالی جدید "+" : "+str(financial_document.id)
            new_log['url']=financial_document.get_absolute_url()
            new_log['description']="سند مالی جدید با موفقیت اضافه گردید."
            self.log(**new_log)
        return result,message,financial_document


            
    def edit_financial_document(self,*args, **kwargs):
        result,message,financial_document=FAILED,"",None
        # if not Permission(request=self.request).is_permitted(APP_NAME,OperationEnum.ADD,"financialdocument"):
        if not self.request.user.has_perm(APP_NAME+".add_financialdocument"):
            message="دسترسی غیر مجاز"
            return result,message,financial_document
         
        financial_document_id=kwargs['financial_document_id']
        financial_document=FinancialDocument.objects.filter(pk=financial_document_id).first()
        if financial_document is None:
            message='سند مالی پیدا نشد.'
            return FAILED,message,None
        
        if 'title' in kwargs:
            financial_document.title=kwargs['title']
        if 'status' in kwargs:
            financial_document.status=kwargs['status']
     
        
        # if 'financial_year_id' in kwargs:
        #     payment.financial_year_id=kwargs['financial_year_id']
        # else:
        #     payment.financial_year_id=FinancialYear.get_by_date(date=payment.transaction_datetime).id
        result,message,financial_document=financial_document.save()
        if result==SUCCEED:
            message="با موفقیت ویرایش شد."
            me_person=PersonRepo(request=self.request).me
            new_log={}
            new_log['title']="ویرایش سند مالی "+" : "+financial_document.title
            new_log['app_name']=APP_NAME
            new_log['url']=financial_document.get_absolute_url()
            new_log['person']=me_person
            new_log['description']="سند مالی جدید با موفقیت ویرایش گردید."
            LogRepo(request=self.request).add_log(**new_log)
        return result,message,financial_document


class FinancialDocumentLineRepo(Repo):
    def __init__(self,request,*args, **kwargs):
        
        super(FinancialDocumentLineRepo,self).__init__(request=request,app_name=APP_NAME,*args,**kwargs)
        if self.me is not None and request.user.has_perm(APP_NAME+".view_financialdocumentline"):
            self.objects=FinancialDocumentLine.objects
        else:
            my_accounts_ids=[]
            for pa in PersonAccount.objects.filter(person_id=self.me.id):
                my_accounts_ids.append(pa.id)
            self.objects=FinancialDocumentLine.objects.filter(Q(account__in=my_accounts_ids))

    def list(self,*args, **kwargs):
        objects=self.objects
        if "start_date" in kwargs :
            start_date=kwargs["start_date"]
            if start_date is not None and not start_date=="":
                objects=objects.filter(date_time__gte=start_date) 
        if "end_date" in kwargs:
            end_date=kwargs["end_date"]
            if end_date is not None and not end_date=="":
                objects=objects.filter(date_time__lte=end_date) 
        if "id__in" in kwargs :
            id__in=kwargs["id__in"]
            objects=objects.filter(id__in=id__in) 
        if "search_for" in kwargs and kwargs["search_for"] is not None and len(kwargs["search_for"])>0 :
            objects=objects.filter(Q(title__contains=kwargs['search_for'])|Q(financial_event__title__contains=kwargs['search_for']) )
        if "amount" in kwargs and kwargs["amount"] is not None and kwargs["amount"]>0 :
            objects=objects.filter(Q(bedehkar=kwargs['amount']) | Q(bestankar=kwargs['amount']) )

            
        if "bestankar" in kwargs and kwargs["bestankar"] is not None:
            objects=objects.filter(Q(bestankar=kwargs['bestankar']) )

            
        if "account_id__in" in kwargs :
            objects=objects.filter(Q(account_id__in=kwargs['account_id__in']) )
        if "bedehkar" in kwargs and kwargs["bedehkar"] is not None:
            objects=objects.filter(Q(bedehkar=kwargs['bedehkar']) )

        if "account_code" in kwargs and kwargs["account_code"] is not None :
            account_code=kwargs["account_code"]
            if not account_code=="":
                account=AccountRepo(request=self.request).account(code=account_code)
                if account is not None:
                    objects=objects.filter(account_id=account.id)
        if "account_id" in kwargs and kwargs["account_id"] is not None and kwargs["account_id"]>0 :
            account_id=kwargs["account_id"]
            objects=objects.filter(account_id=account_id)
        if "event_id" in kwargs and kwargs["event_id"] is not None and kwargs["event_id"]>0 :
            event_id=kwargs["event_id"]
            objects=objects.filter(financial_event_id=event_id)

        if "financial_event_id" in kwargs and kwargs["financial_event_id"] is not None and kwargs["financial_event_id"]>0 :
            financial_event_id=kwargs["financial_event_id"]
            objects=objects.filter(financial_event_id=financial_event_id)
        
        if "financial_document_id" in kwargs and kwargs["financial_document_id"] is not None and kwargs["financial_document_id"]>0 :
            financial_document_id=kwargs["financial_document_id"]
            objects=objects.filter(financial_document_id=financial_document_id)
        
        return objects.all().order_by('date_time')

    def financial_document_line(self,*args, **kwargs):
        if "financial_document_line_id" in kwargs and kwargs["financial_document_line_id"] is not None:
            return self.objects.filter(pk=kwargs['financial_document_line_id']).first() 
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        
    def add_financial_document_line(self,*args, **kwargs):
        financial_document_line,message,result=(None,"",FAILED)
    
        bestankar=kwargs['bestankar']
        bedehkar=kwargs['bedehkar']
        if bedehkar==0 and bestankar==0:
            message="مبلغ بدهکار و بستانکار صفر وارد شده است."
            return result,message,financial_document_line

            
        if bedehkar>0 and bestankar>0:
            message="مبلغ بدهکار و بستانکار ، هر دو وارد شده است."
            return result,message,financial_document_line

        if bedehkar<0 or bestankar<0:
            message="مبلغ بدهکار یا بستانکار منفی وارد شده است."
            return result,message,financial_document_line

        # if not Permission(request=self.request).is_permitted(APP_NAME,OperationEnum.ADD,"accountingdocumentline"):
        if not self.request.user.has_perm(APP_NAME+".add_accountingdocumentline"):
            message="دسترسی غیر مجاز"
            return result,message,financial_document_line
        
        financial_document_line=FinancialDocumentLine()
        if 'title' in kwargs:
            financial_document_line.title=kwargs['title']
        if 'status' in kwargs:
            financial_document_line.status=kwargs['status']
        if 'financial_event_id' in kwargs:
            financial_event_id=kwargs['financial_event_id']
            financial_event=FinancialEvent.objects.filter(pk=financial_event_id).first()
            if financial_event is None:
                message='رویداد مالی درست انتخاب نشده است.'
                return result,message,None
            financial_document_line.financial_event_id=financial_event_id
        if 'financial_document_id' in kwargs:
            financial_document_id=kwargs['financial_document_id']
            if int(financial_document_id)==0 and 'financial_document_title' in kwargs:
                
                result,message,financial_document=FinancialDocumentRepo(request=self.request).add_financial_document(title=kwargs['financial_document_title'])
                if financial_document is not None:
                    financial_document_id=financial_document.id
                else:
                    # message='سند مالی جدید ایجاد نشد.'
                    return result,message,None
            
            financial_document_line.financial_document_id=financial_document_id
            financial_document=financial_document_line.financial_document
            if financial_document.status==FinancialDocumentStatusEnum.ACCEPTED:
                message='سند مرتبط تایید شده می باشد.<br> نمی توان به سطر های آن افزود'
                return FAILED,message,None
        if 'description' in kwargs:
            financial_document_line.description=kwargs['description']
        if 'persian_date_time' in kwargs and kwargs['persian_date_time'] is not None and not kwargs['persian_date_time']=='':
            persian_date_time=kwargs['persian_date_time']
            date_time=PersianCalendar().to_gregorian(persian_date_time)
            # date_time=date_time,persian_date_time=kwargs['persian_date_time'])
            # financial_document_line.date_time=date_time
        if 'bestankar' in kwargs  :
            financial_document_line.bestankar=kwargs['bestankar']
        if 'bedehkar' in kwargs :
            financial_document_line.bedehkar=kwargs['bedehkar'] 
        if 'date_time' in kwargs :

            date_time=kwargs['date_time']
            year=date_time[:2]
            if year=="13" or year=="14":
                date_time=PersianCalendar().to_gregorian(kwargs["date_time"])
            financial_document_line.date_time=date_time 

        if 'account_code' in kwargs and kwargs['account_code'] is not None:
            account=AccountRepo(request=self.request).account(code=kwargs['account_code']) 
            if account is not None:
                financial_document_line.account=account
        if 'account_id' in kwargs and kwargs['account_id'] is not None:
            financial_document_line.account_id=kwargs['account_id'] 
        
        
        # if 'financial_year_id' in kwargs:
        #     payment.financial_year_id=kwargs['financial_year_id']
        # else:
        #     payment.financial_year_id=FinancialYear.get_by_date(date=payment.transaction_datetime).id

        if financial_document_line.account.nature==AccountNatureEnum.ONLY_BESTANKAR and financial_document_line.bedehkar>0:
            message=financial_document_line.account.name+" ماهیت فقط بستانکار دارد"
            financial_document_line=None
            return result,message,financial_document_line
        if financial_document_line.account.nature==AccountNatureEnum.ONLY_BEDEHKAR and financial_document_line.bestankar>0:
            message=financial_document_line.account.name+" ماهیت فقط بدهکار دارد"
            financial_document_line=None
            return result,message,financial_document_line

        result,message,financial_document_line=financial_document_line.save()
        if result==FAILED:
            return result,message,financial_document_line 
        result=SUCCEED
        message="با موفقیت اضافه گردید."
         

        new_log={}
        new_log['title']="سطر سند مالی جدید "+" "+str(financial_document_line.id)
        new_log['url']=financial_document_line.get_absolute_url()
        new_log['description']="خط سطر سند مالی جدید با موفقیت اضافه گردید."
        # LogRepo(request=self.request).add_log(**new_log)
        self.log(**new_log)
        return result,message,financial_document_line
    

       
    def edit_financial_document_line(self,*args, **kwargs):
        desc=""
        financial_document_line,message,result=(None,"",FAILED)
    
        bestankar=kwargs['bestankar']
        bedehkar=kwargs['bedehkar']
        if bedehkar==0 and bestankar==0:
            message="مبلغ بدهکار و بستانکار صفر وارد شده است."
            return result,message,financial_document_line

            
        if bedehkar>0 and bestankar>0:
            message="مبلغ بدهکار و بستانکار ، هر دو وارد شده است."
            return result,message,financial_document_line

        if bedehkar<0 or bestankar<0:
            message="مبلغ بدهکار یا بستانکار منفی وارد شده است."
            return result,message,financial_document_line

        # if not Permission(request=self.request).is_permitted(APP_NAME,OperationEnum.ADD,"accountingdocumentline"):
        if not self.request.user.has_perm(APP_NAME+".change_accountingdocumentline"):
            message="دسترسی غیر مجاز"
            return result,message,financial_document_line
        
        financial_document_line=FinancialDocumentLine.objects.filter(pk=kwargs['financial_document_line_id']).first()
        if financial_document_line is None:
            return FAILED,'سطر پیدا نشد.',None
        
        if 'status' in kwargs:
            financial_document_line.status=kwargs['status']
            
        if 'title' in kwargs:
            desc=financial_document_line.title
            financial_document_line.title=kwargs['title']
        if 'financial_event_id' in kwargs:
            financial_event_id=kwargs['financial_event_id']
            financial_event=FinancialEvent.objects.filter(pk=financial_event_id).first()
            if financial_event is None:
                message='رویداد مالی درست انتخاب نشده است.'
                return result,message,None
            financial_document_line.financial_event_id=financial_event_id
        if 'financial_document_id' in kwargs:
            financial_document_id=kwargs['financial_document_id']
            financial_document=FinancialDocument.objects.filter(pk=financial_document_id).first()
            if financial_document is None:
                message='سند مالی درست انتخاب نشده است.'
                return FAILED,message,None
 
            if financial_document.status==FinancialDocumentStatusEnum.ACCEPTED:
                message='سند مرتبط تایید شده می باشد.<br> نمی توان سطر های آن را تغییر داد.'
                return FAILED,message,None
            financial_document_line.financial_document_id=financial_document_id
        if 'description' in kwargs:
            financial_document_line.description=kwargs['description']
        if 'persian_date_time' in kwargs and kwargs['persian_date_time'] is not None and not kwargs['persian_date_time']=='':
            persian_date_time=kwargs['persian_date_time']
            date_time=PersianCalendar().to_gregorian(persian_date_time)
            # date_time=date_time,persian_date_time=kwargs['persian_date_time'])
            # financial_document_line.date_time=date_time
        if 'bestankar' in kwargs  :
            financial_document_line.bestankar=kwargs['bestankar']
        if 'bedehkar' in kwargs :
            financial_document_line.bedehkar=kwargs['bedehkar'] 
        if 'date_time' in kwargs :

            date_time=kwargs['date_time']
            year=date_time[:2]
            if year=="13" or year=="14":
                date_time=PersianCalendar().to_gregorian(kwargs["date_time"])
            financial_document_line.date_time=date_time 

        if 'account_code' in kwargs and kwargs['account_code'] is not None:
            account=AccountRepo(request=self.request).account(code=kwargs['account_code']) 
            if account is not None:
                financial_document_line.account=account
        if 'account_id' in kwargs and kwargs['account_id'] is not None:
            financial_document_line.account_id=kwargs['account_id'] 
        
        if financial_document_line.account is None:
            message='حساب درست انتخاب نشده است.'
            return FAILED,message,None
        # if 'financial_year_id' in kwargs:
        #     payment.financial_year_id=kwargs['financial_year_id']
        # else:
        #     payment.financial_year_id=FinancialYear.get_by_date(date=payment.transaction_datetime).id

        if financial_document_line.account.nature==AccountNatureEnum.ONLY_BESTANKAR and financial_document_line.bedehkar>0:
            message=financial_document_line.account.name+" ماهیت فقط بستانکار دارد"
            financial_document_line=None
            return result,message,financial_document_line
        if financial_document_line.account.nature==AccountNatureEnum.ONLY_BEDEHKAR and financial_document_line.bestankar>0:
            message=financial_document_line.account.name+" ماهیت فقط بدهکار دارد"
            financial_document_line=None
            return result,message,financial_document_line

        result,message,financial_document_line=financial_document_line.save()
        if result==FAILED:
            return result,message,financial_document_line 
        result=SUCCEED
        message="با موفقیت تغییر یافت."
         

        me_person=PersonRepo(request=self.request).me
   
        
        new_log={}
        new_log['title']="ویرایش سطر سند مالی  "+" "+str(financial_document_line.id)
        new_log['url']=financial_document_line.get_absolute_url()
        new_log['description']="خط سطر سند مالی جدید با موفقیت ویرایش گردید."+"عنوان قبلی : "+desc
        # LogRepo(request=self.request).add_log(**new_log)
        self.log(**new_log)

        return result,message,financial_document_line
    

    def delete_all(self,*args,**kwargs):
        result,message=FAILED,''
        if not self.request.user.has_perm(APP_NAME+".delete_accountingdocumentline"):
            message="دسترسی غیر مجاز"
            return result,message
        FinancialDocumentLine.objects.all().delete() 
                   
        result=SUCCEED
        message="همه حذف شدند."
        return result,message
    
    
    
    def add_event_financial_document_line(self,*args, **kwargs):
        result,message,financial_document_line=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_financialdocumentline"):
            message="دسترسی غیر مجاز"
            return result,message,accounting_document_line

        event_id=kwargs['event_id']
        bestankar=kwargs['bestankar']
        bedehkar=kwargs['bedehkar']
        account_code=kwargs['account_code']
        financial_document_id=kwargs['financial_document_id']
        financial_document_title=kwargs['financial_document_title']

        account_repo=AccountRepo(request=self.request)
        event_repo=FinancialEventRepo(request=self.request)
        financial_document_repo=FinancialDocumentRepo(request=self.request)

        event=event_repo.event(pk=event_id)
        account=account_repo.account(code=account_code)

        if financial_document_id==0:
            result,message,financial_document=financial_document_repo.add_financial_document(title=financial_document_title)
        else:
            financial_document=financial_document_repo.financial_document(pk=financial_document_id)
        if account is not None and financial_document is not None and event is not None:
            financial_document_line=FinancialDocumentLine()
            financial_document_line.event=event
            financial_document_line.bestankar=bestankar
            financial_document_line.bedehkar=bedehkar
            financial_document_line.account=account
            financial_document_line.date_time=event.event_datetime
            financial_document_line.title=event.title
            financial_document_line.financial_document=financial_document
            financial_document_line.save()
            result=SUCCEED
            message="اضافه شد."
            return result,message,financial_document_line
  

class FinancialEventRepo(Repo):
    def __init__(self,request,*args, **kwargs):
        super(FinancialEventRepo,self).__init__(request,app_name=APP_NAME,*args, **kwargs)
        self.my_financial_events=[]
        self.objects=FinancialEvent.objects.filter(id=0)


        if self.me_person is not None and request.user.has_perm(APP_NAME+".view_financialevent"):
            self.objects=FinancialEvent.objects 
        elif self.me_person is not None:
            my_accounts=AccountRepo(request=request).my_accounts
            ids=[]
            for acc in my_accounts:
                ids.append(acc.id)
            
             
            self.my_financial_events=FinancialEvent.objects.filter(Q(bedehkar_id__in=ids)|Q(bestankar_id__in=ids))
            self.objects=self.my_financial_events
    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            objects=objects.filter(Q(title__contains=search_for)  )
        if "raw" in kwargs:
            raw=kwargs["raw"]
            if raw:
                objects=objects.filter(pk=0)
        if "account_code" in kwargs:
            account_code=kwargs["account_code"]
            objects=objects.filter(Q(bedehkar__code=account_code) | Q(bestankar__code=account_code)  )
        if "account_id" in kwargs:
            account_id=kwargs["account_id"]
            objects=objects.filter(Q(bedehkar_id=account_id) | Q(bestankar_id=account_id)  )
        if "parent_id" in kwargs:
            parent_id=kwargs["parent_id"]
            objects=objects.filter(parent_id=parent_id)  
        return objects.order_by('-event_datetime')
       
    def roots(self,*args, **kwargs):
        objects=self.objects.filter(parent_id=None)
        return objects.all()

    def financial_event(self,*args, **kwargs):
        if "financial_event_id" in kwargs and kwargs["financial_event_id"] is not None:
            return self.objects.filter(pk=kwargs['financial_event_id']).first()  
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "event_id" in kwargs and kwargs["event_id"] is not None:
            return self.objects.filter(pk=kwargs['event_id']).first() 
         
    def edit_financial_event(self,*args, **kwargs):
        result,message,financial_event=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_financialevent"):
            message="دسترسی غیر مجاز"
            return result,message,financial_event
    
        if "financial_event_id" in kwargs and kwargs["financial_event_id"] is not None:
            financial_event=FinancialEvent.objects.filter(pk=kwargs['financial_event_id']).first()  
            if financial_event is None:
                message="رویداد پیدا نشد."
                return result,message,financial_event

        if "title" in kwargs and kwargs["title"] is not None:
            financial_event.title=kwargs['title']


        if "status" in kwargs  :
            financial_event.status=kwargs['status']


        if "amount" in kwargs and kwargs["amount"] is not None:
            financial_event.amount=kwargs['amount']

        if "payment_method" in kwargs and kwargs["payment_method"] is not None:
            financial_event.payment_method=kwargs['payment_method']

            
        if "shipping_fee" in kwargs and kwargs["shipping_fee"] is not None:
            financial_event.shipping_fee=kwargs['shipping_fee'] 

        if "discount" in kwargs and kwargs["discount"] is not None:
            financial_event.discount=kwargs['discount']
        if "tax_percentage" in kwargs and kwargs["tax_percentage"] is not None:
            financial_event.tax_percentage=kwargs['tax_percentage']

        if "shipping_fee" in kwargs and kwargs["shipping_fee"] is not None:
            financial_event.shipping_fee=kwargs['shipping_fee']

        if "bedehkar_id" in kwargs and kwargs["bedehkar_id"] is not None:
            financial_event.bedehkar_id=kwargs['bedehkar_id']

        if "bestankar_id" in kwargs and kwargs["bestankar_id"] is not None:
            financial_event.bestankar_id=kwargs['bestankar_id']
        
        if 'event_datetime1' in kwargs and kwargs['event_datetime'] is not None:
            event_datetime=kwargs["event_datetime"]
            financial_event.event_datetime=event_datetime

           
            year=event_datetime[:2]
            if year=="13" or year=="14":
                event_datetime=PersianCalendar().to_gregorian(event_datetime)
            financial_event.event_datetime=event_datetime

        # if 'start_datetime' in kwargs:
        #     start_datetime=kwargs["start_datetime"]
        #     financial_event.start_datetime=start_datetime

           
        #     year=start_datetime[:2]
        #     if year=="13" or year=="14":
        #         start_datetime=PersianCalendar().to_gregorian(start_datetime)
        #     financial_event.start_datetime=start_datetime

        # if 'end_datetime' in kwargs:
        #     end_datetime=kwargs["end_datetime"]
        #     financial_event.end_datetime=end_datetime

           
        #     year=end_datetime[:2]
        #     if year=="13" or year=="14":
        #         end_datetime=PersianCalendar().to_gregorian(end_datetime)
        #     financial_event.end_datetime=end_datetime

        result,message,financial_event=financial_event.save()
        
            # if 'title' in kwargs:            
            #     log.title=kwargs['title']
            # if 'url' in kwargs:     
            #     log.url=kwargs['url']
            # if 'app_name' in kwargs:     
            #     log.app_name=kwargs['app_name']
            # if 'person' in kwargs:            
            #     log.person=kwargs['person'] 
            # if 'person_id' in kwargs:            
            #     log.person_id=kwargs['person_id']
            # if 'description' in kwargs:            
            #     log.description=kwargs['description']
        kwargs['title']='ویرایش رویداد مالی'    
        kwargs['url']=financial_event.get_absolute_url()    
        kwargs['description']='description'    
        self.log(**kwargs)
        return result,message,financial_event
       
    def delete_all(self,*args, **kwargs):
        result,message=FAILED,""
        if not self.request.user.has_perm(APP_NAME+".delete_financialevent"):
            message="دسترسی غیر مجاز"
            return result,message
        financial_events=FinancialEvent.objects.all()
        financial_events.delete()
        result=SUCCEED
        message='همه رویداد ها حذف شدند.'
        return result,message
    
    def add_financial_event(self,*args,**kwargs):
        result,message,financial_event=FAILED,"",None
        
        # if len(FinancialEvent.objects.filter(title=kwargs['title']))>0:
        #     message='عنوان تکراری' 
        #     return result,message,None
        
        if not self.request.user.has_perm(APP_NAME+".add_financialevent"):
            message="دسترسی غیر مجاز"
            return result,message,financial_event

        financial_event=FinancialEvent()
        if 'bedehkar_id' in kwargs:
            financial_event.bedehkar_id=kwargs["bedehkar_id"]
        if 'bestankar_id' in kwargs:
            financial_event.bestankar_id=kwargs["bestankar_id"]

        if 'valid' in kwargs and kwargs['valid'] is not None:
            financial_event.valid=kwargs["valid"]

            
        if "shipping_fee" in kwargs and kwargs["shipping_fee"] is not None:
            financial_event.shipping_fee=kwargs['shipping_fee'] 

        if "payment_method" in kwargs and kwargs["payment_method"] is not None:
            financial_event.payment_method=kwargs['payment_method'] 

        if "status" in kwargs and kwargs["status"] is not None:
            financial_event.status=kwargs['status'] 

        if 'title' in kwargs and kwargs["title"] is not None:
            financial_event.title=kwargs["title"]
        if 'description' in kwargs and kwargs["description"] is not None:
            financial_event.description=kwargs["description"]
        if 'parent_id' in kwargs and kwargs["parent_id"] is not None:
            if kwargs["parent_id"]>0:
                financial_event.parent_id=kwargs["parent_id"]
        if 'color' in kwargs and kwargs["color"] is not None:
            financial_event.color=kwargs["color"]
        if 'amount' in kwargs and kwargs["amount"] is not None:
            financial_event.amount=kwargs["amount"]
        if 'priority' in kwargs and kwargs["priority"] is not None:
            financial_event.priority=kwargs["priority"]
        if 'type' in kwargs and kwargs["type"] is not None:
            financial_event.type=kwargs["type"]
        if 'event_datetime' in kwargs and kwargs["event_datetime"] is not None:
            event_datetime=kwargs["event_datetime"]
            financial_event.event_datetime=event_datetime

           
            year=event_datetime[:2]
            if year=="13" or year=="14":
                event_datetime=PersianCalendar().to_gregorian(event_datetime)
            financial_event.event_datetime=event_datetime

        # if 'start_datetime' in kwargs:
        #     start_datetime=kwargs["start_datetime"]
        #     financial_event.start_datetime=start_datetime

           
        #     year=start_datetime[:2]
        #     if year=="13" or year=="14":
        #         start_datetime=PersianCalendar().to_gregorian(start_datetime)
        #     financial_event.start_datetime=start_datetime

        # if 'end_datetime' in kwargs:
        #     end_datetime=kwargs["end_datetime"]
        #     financial_event.end_datetime=end_datetime

           
        #     year=end_datetime[:2]
        #     if year=="13" or year=="14":
        #         end_datetime=PersianCalendar().to_gregorian(end_datetime)
        #     financial_event.end_datetime=end_datetime

        (result,message,financial_event)=financial_event.save()
        kwargs['title']='افزودن رویداد مالی'    +" "+str(financial_event.id)
        kwargs['url']=financial_event.get_absolute_url()    
        kwargs['description']='description'    
        self.log(**kwargs)

        return result,message,financial_event
 

class InvoiceRepo(FinancialEventRepo):
    def __init__(self,request,*args, **kwargs):
        self.me=None
        super().__init__(request,*args, **kwargs)
        self.my_invoices=Invoice.objects.filter(pk=0)
        self.request=request
        self.objects=self.my_invoices
        person=PersonRepo(request=request).me
        if request.user.has_perm(APP_NAME+".view_invoice"):
            self.objects=Invoice.objects
        elif person is not None:
            ids=[]
            for financial_event in self.my_financial_events:
                ids.append(financial_event.id)
            
            self.my_invoices=Invoice.objects.filter(id__in=ids)
            self.objects=self.my_invoices
    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            objects=objects.filter(Q(title__contains=search_for)  )
        if "parent_id" in kwargs:
            parent_id=kwargs["parent_id"]
            objects=objects.filter(parent_id=parent_id)  
        return objects.all()
       
    def roots(self,*args, **kwargs):
        objects=self.objects.filter(parent_id=None)
        return objects.all()

    def invoice(self,*args, **kwargs):
        if "invoice_id" in kwargs and kwargs["invoice_id"] is not None:
            return self.objects.filter(pk=kwargs['invoice_id']).first()  
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
         
        if "invoice_no" in kwargs and kwargs["invoice_no"] is not None:
            return self.objects.filter(invoice_no=kwargs['invoice_no']).first() 
         
       

    def add_invoice(self,*args,**kwargs):
        result,message,invoice=FAILED,"",None
        
        if len(Invoice.objects.filter(title=kwargs['title']))>0:
            message='عنوان تکراری' 
            return result,message,invoice
            
        if not self.request.user.has_perm(APP_NAME+".add_invoice"):
            message="دسترسی غیر مجاز"
            return result,message,invoice

        invoice=Invoice()
        
        if 'valid' in kwargs and kwargs['valid'] is not None:
            invoice.valid=kwargs["valid"]

        if 'title' in kwargs:
            invoice.title=kwargs["title"]
        if 'parent_id' in kwargs:
            if kwargs["parent_id"]>0:
                invoice.parent_id=kwargs["parent_id"]
        if 'color' in kwargs:
            invoice.color=kwargs["color"]
        if 'code' in kwargs:
            invoice.code=kwargs["code"]
        if 'priority' in kwargs:
            invoice.priority=kwargs["priority"]
        if 'bedehkar_id' in kwargs:
            invoice.bedehkar_id=kwargs["bedehkar_id"]
        if 'bestankar_id' in kwargs:
            invoice.bestankar_id=kwargs["bestankar_id"]
        if 'event_datetime' in kwargs:
            
            year=kwargs['event_datetime'][:2]
            if year=="13" or year=="14":
                kwargs['event_datetime']=PersianCalendar().to_gregorian(kwargs["event_datetime"])
            invoice.event_datetime=kwargs["event_datetime"]

        if 'type' in kwargs:
            invoice.type=kwargs["type"]

           
        if 'status' in kwargs:
            invoice.status=kwargs["status"]

           
           
        if 'invoice_no' in kwargs:
            invoice.invoice_no=kwargs["invoice_no"]

           
        (result,message,invoice)=invoice.save()
        return result,message,invoice

    def edit_invoice(self,*args, **kwargs):
        result,message,invoice=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_invoice"):
            message="دسترسی غیر مجاز"
            return result,message,invoice

        invoice=Invoice.objects.filter(pk=kwargs['invoice_id']).first()

         
        if invoice.status==FinancialEventStatusEnum.APPROVED:
            message='فاکتور تایید شده و امکان تغییر ، ویرایش و افزودن سطر وجود ندارد.'
            return FAILED,message,None
        
        if invoice.status==FinancialEventStatusEnum.DELIVERED:
            message='فاکتور تحویل شده و امکان تغییر ، ویرایش و افزودن سطر وجود ندارد.'
            return FAILED,message,None
        
        if invoice.status==FinancialEventStatusEnum.FINISHED:
            message='فاکتور نهایی شده و امکان تغییر ، ویرایش و افزودن سطر وجود ندارد.'
            return FAILED,message,None
    
        if invoice is None:
            message="فاکتور پیدا نشد."
            return result,message,invoice
        if 'title' in kwargs:
            invoice.title=kwargs['title'] 



        if 'valid' in kwargs and kwargs['valid'] is not None:
            invoice.valid=kwargs['valid'] 


        if 'bedehkar_id' in kwargs and kwargs['bedehkar_id'] is not None and not kwargs['bedehkar_id']=='':
            invoice.bedehkar_id=kwargs['bedehkar_id']
            
        if 'discount' in kwargs and kwargs['discount'] is not None and not kwargs['discount']=='':
            invoice.discount=kwargs['discount'] 

        if 'tax_percentage' in kwargs and kwargs['tax_percentage'] is not None and not kwargs['tax_percentage']=='':
            invoice.tax_percentage=kwargs['tax_percentage'] 
            
        if 'shipping_fee' in kwargs and kwargs['shipping_fee'] is not None:
            invoice.shipping_fee=kwargs['shipping_fee']
            
        if 'status' in kwargs and kwargs['status'] is not None and not kwargs['status']=='':
            invoice.status=kwargs['status']
            
            
        if 'event_datetime' in kwargs and kwargs['event_datetime'] is not None and not kwargs['event_datetime']=='':
            year=kwargs['event_datetime'][:2]
            if year=="13" or year=="14":
                kwargs['event_datetime']=PersianCalendar().to_gregorian(kwargs["event_datetime"])
            invoice.event_datetime=kwargs["event_datetime"]
        if 'payment_method' in kwargs and kwargs['payment_method'] is not None and not kwargs['payment_method']=='':
            invoice.payment_method=kwargs['payment_method']

            
        if 'description' in kwargs  :
            invoice.description=kwargs['description']   
            
        if 'short_description' in kwargs  :
            invoice.short_description=kwargs['short_description']   

        if 'bestankar_id' in kwargs and kwargs['bestankar_id'] is not None:
            invoice.bestankar_id=kwargs['bestankar_id']   

        if 'invoice_lines' in kwargs and kwargs['invoice_lines'] is not None and not kwargs['invoice_lines']=='':
            invoice_lines=kwargs['invoice_lines']   
            for new_invoice_line in invoice_lines:
                invoice_line=InvoiceLineRepo(request=self.request).invoice_line(pk=int(new_invoice_line['invoice_line_id']))
                if invoice_line is not None:
                    invoice_line.row=int(new_invoice_line['row'])
                    invoice_line.quantity=float(new_invoice_line['quantity'])
                    invoice_line.unit_name=new_invoice_line['unit_name']
                    invoice_line.unit_price=int(new_invoice_line['unit_price'])
                    invoice_line.discount_percentage=int(new_invoice_line['discount_percentage'])
                    invoice_line.save(normalize_row=False)
                    if invoice_line.quantity==0:
                        invoice_line.delete()

        return invoice.save()


class CategoryRepo():

    def __init__(self,request,*args, **kwargs):
        self.me=None
        self.my_categorys=[]
        self.request=request
        self.objects=Category.objects.filter(id=0)
        person=PersonRepo(request=request).me
         
        self.objects=Category.objects.order_by('priority')

    def delete_all(self):
        
        if self.request.user.has_perm(APP_NAME+".delete_category"):
            Category.objects.all().delete()
            result=SUCCEED
            message="همه دسته بندی های کالاها با موفقیت حذف شدند."
        return result,message


    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            objects=objects.filter(Q(title__contains=search_for))
        if "parent_id" in kwargs:
            parent_id=kwargs["parent_id"]
            objects=objects.filter(parent_id=parent_id)  

        if "for_home" in kwargs:
            for_home=kwargs["for_home"]
            if for_home:
                objects=objects.filter(Q(parent_id=None) | Q(priority=0))
        return objects.all()
       
    def roots(self,*args, **kwargs):
        objects=self.objects.filter(parent_id=None)
        return objects.all()

    def category(self,*args, **kwargs):
        if "category_id" in kwargs and kwargs["category_id"] is not None:
            return self.objects.filter(pk=kwargs['category_id']).first()  
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "code" in kwargs and kwargs["code"] is not None:
            code=kwargs['code']
            code=kwargs['code']

            category= self.objects.filter(code=code).first()
            return category
        if "category_code" in kwargs and kwargs["category_code"] is not None:
            a= self.objects.filter(code=kwargs['category_code']).first() 
            if a is not None:
                return a
            else:
                try:
                    a= self.objects.filter(pure_code=filter_number(kwargs['category_code'])).first() 
                    if a is not None:
                        return a
                except:
                    pass
          
    def set_priority(self,*args, **kwargs):
        result,message,priority=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".change_category"):
            return result,message,category_tags
        priority=kwargs['priority']
        category_id=kwargs['category_id']
        category=Category.objects.filter(pk=category_id).first()
        if category is not None:
            category.priority=priority
            category.save()
        result=SUCCEED
        return result,message,priority

    def set_category_parent(self,*args, **kwargs):
        result,message,category,parent=FAILED,"",None,None
        category=self.category(*args,**kwargs)
        parent=None
        parent_code=kwargs["parent_code"]
        parent=self.category(category_code=parent_code)
        category.parent=parent
        category.save()
        result=SUCCEED
        message="با موفقیت تغییر یافت"
        return result,message,category,parent

    def initial_default_categories(self,*args, **kwargs):
        category_group_counter=0
        basic_categorys_counter=0
        moein_categorys_counter=0
        moein2_categorys_counter=0
        tafsili_categorys_counter=0 

        result=SUCCEED
        message=""  
        if not self.request.user.has_perm(APP_NAME+".add_category"):
            message="دسترسی غیر مجاز"
            result=FAILED
            return message,result
        
        categorys=default_categorys() 

        for category in categorys:
            parent_category=None
            if 'parent_code' in category:
                parent_category=Category.objects.filter(code=category["parent_code"]).first()
            new_category=Category(name=category["name"],color=category["color"],code=category['code'],priority=category['priority'],parent=parent_category)
            # new_category=Category(parent=parent_category,**kwargs)
            new_category.save()
            # category_group_counter+=1
            # basic_categorys_counter+=1   
 
                                  
        categorys=Category.objects
        category_group_counter=len(categorys.filter(type=CategoryTypeEnum.GROUP))
        basic_categorys_counter=len(categorys.filter(type=CategoryTypeEnum.BASIC))
        moein_categorys_counter=len(categorys.filter(type=CategoryTypeEnum.MOEIN_1))
        moein2_categorys_counter=len(categorys.filter(type=CategoryTypeEnum.MOEIN_2))
        tafsili_categorys_counter=len(categorys.filter(Q(type=CategoryTypeEnum.TAFSILI_1)|Q(type=CategoryTypeEnum.TAFSILI_2)|Q(type=CategoryTypeEnum.TAFSILI_3)|Q(type=CategoryTypeEnum.TAFSILI_4)))
                                                        

        if result==SUCCEED:
            message="با موفقیت اضافه گردید."
        message+=f"<br>{category_group_counter}   گروه حساب" 
        message+=f"<br>{basic_categorys_counter}   حساب  کل " 
        message+=f"<br>{moein_categorys_counter}  حساب معین سطح یک " 
        message+=f"<br>{moein2_categorys_counter}  حساب معین سطح دو " 
        message+=f"<br>{tafsili_categorys_counter}  حساب تفصیلی " 

        me_person=PersonRepo(request=self.request).me
        new_log={}
        new_log['title']="افزودن حساب های پیش فرض"
        new_log['app_name']=APP_NAME
        new_log['person']=me_person
        new_log['description']="حساب های پیش فرض اضافه شدند."
        LogRepo(request=self.request).add_log(**new_log)
        return result,message
 
    def add_category(self,*args,**kwargs):
        result,message,category=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_category"):
            message="دسترسی غیر مجاز"
            return result,message,category
        category=Category()
        if 'title' in kwargs:
            category.title=kwargs["title"]
        if 'parent_id' in kwargs :
            if kwargs["parent_id"] is not None and kwargs["parent_id"]>0:
                category.parent_id=kwargs["parent_id"]
        if 'color' in kwargs:
            category.color=kwargs["color"]
        if 'id' in kwargs:
            category.id=kwargs["id"]
        if 'priority' in kwargs and kwargs['priority'] is not None:
            category.priority=kwargs["priority"]
        if 'thumbnail_origin' in kwargs and kwargs['thumbnail_origin'] is not None:
            category.thumbnail_origin=kwargs["thumbnail_origin"]
        
        (result,message,category)=category.save()

        return result,message,category
    
    def add_product_to_category(self,*args, **kwargs):
        result,message,product_categories,product,category=FAILED,'',[],None,None
            
        if not self.request.user.has_perm(APP_NAME+".add_category"):
            message="دسترسی غیر مجاز"
            return result,message,product_categories,product,category
        # product_id=0
        # category_id=0
        # if 'category_id' in kwargs:
        #     category_id=kwargs["category_id"]
        # if 'product_id' in kwargs:
        #     product_id=kwargs["product_id"]
        product=ProductRepo(request=self.request).product(*args, **kwargs)
        category=self.category(*args, **kwargs)
        if product is None:
            message="کالایی پیدا نشد"
            return result,message,product_categories,product,category
        if category is None:
            message="دسته بندی پیدا نشد"
            return result,message,product_categories,product,category
        if product in category.products.all():
            message='با موفقیت کالا از این دسته بندی حذف شد.'
            result=SUCCEED
            category.products.remove(product.id)
            product_categories=product.category_set.all()
            return result,message,product_categories,product,category
        category.products.add(product.id)
        result=SUCCEED
        message='با موفقیت کالا به دسته بندی اضافه شد.'
        product_categories=product.category_set.all()
        return result,message,product_categories,product,category
    
    def import_categories_from_excel(self,*args,**kwargs):
        result,message,categories=FAILED,"",[]
        imported_ids=[]
        modified=added=0 
        
        excel_file=None
        if 'excel_file' in kwargs:
            excel_file=kwargs['excel_file']
        parent_id=0
        if 'parent_id' in kwargs:
            parent_id=kwargs['parent_id']
        categories_to_import=None
        if 'categories_to_import' in kwargs:
            categories_to_import=kwargs['categories_to_import']
        # import pandas
        
        # df = pandas.read_excel(excel_file)
        # categories=[]
        # for row in df.columns[0]:
        #     print (df.columns)
        import openpyxl 
        if excel_file is not None:
            wb = openpyxl.load_workbook(excel_file)
            try:
                ws = wb['categories']
            except:
                message='فایل شما برگه دسته بندی ها ندارد.'
                return result,message,None
            count=kwargs['count']
            try:
                count=int(ws.cell(row=1, column=2).value)
            except:
                message='فایل برگه دسته بندی ها ، تعداد ندارد.'
                return result,message,None 

            categories_to_import=[]
            START_ROW=EXCEL_PRODUCTS_DATA_START_ROW

            for i in range(START_ROW,count+START_ROW):
                category={}
                i=str(i) 
                iiiddd=ws['B'+i].value
                if iiiddd is not None:
                    category['id']=int(ws['B'+i].value)
                    category['new_id']=None
                    category['title']=(ws['C'+i].value)
                    category['parent_id']=(ws['D'+i].value)
                    category['new_parent_id']=None
                    category['thumbnail_origin']=(ws['E'+i].value)
                    # category['id']=0
                    if category['title'] is not None and not category['title']=="":
                        categories_to_import.append(category) 
        # self.import_categories(categories_to_import=categories_to_import,parent_id=0) 
        counter=0
        length=len(categories_to_import)
        imported_ids=[]
        while counter<length:
            for cat in categories_to_import:
                if cat['id'] not in imported_ids:
                    parent_id=cat['parent_id']
                    if parent_id==0:
                        parent_id=None

                    # parent_id=None
                    if parent_id is not None and Category.objects.filter(pk=cat['parent_id']).first() is None:
                        continue
                    result,message,new_category=self.add_category(id=cat['id'],title=cat['title'],parent_id=parent_id,thumbnail_origin=cat['thumbnail_origin'])
                    if result==SUCCEED:
                        counter+=1
                        imported_ids.append(cat['id'])
                    
        result=SUCCEED
        message=f"""{added} دسته بندی اضافه شد.
                    <br>
                    {modified} دسته بندی ویرایش شد. """
        categories=self.list()

        if True:
            log_data={}
            from log.repo import LogRepo
            log_data['person_id']=PersonRepo(request=self.request).me.id
            log_data['url']=reverse("accounting:categories")
            log_data['title']="بازیابی دسته بندی ها"
            log_data['description']=message
            log_data['app_name']=APP_NAME
            LogRepo(request=self.request).add_log(**log_data)


        return result,message,categories


class ChequeRepo():
    def __init__(self,request,*args, **kwargs):
        self.me=None
        self.my_cheques=[]
        self.request=request
        self.objects=Cheque.objects.filter(id=0)
        person=PersonRepo(request=request).me


        if request.user.has_perm(APP_NAME+".view_financialevent"):
            self.objects=Cheque.objects 
        elif person is not None:
            my_accounts=AccountRepo(request=request).my_accounts
            ids=[]
            for acc in my_accounts:
                ids.append(acc.id)
            
             
            self.my_cheques=Cheque.objects.filter(Q(bedehkar_id__in=ids)|Q(bestankar_id__in=ids))
            self.objects=self.my_cheques
    
    def change_image(self,cheque_id,image):
        result,message,cheque=FAILED,'',None
        cheque=self.cheque(cheque_id=cheque_id)
        if cheque is not None:
            cheque.image_origin = image
            cheque.save()
            return SUCCEED,message,cheque
        return FAILED,message,cheque
    
    def list(self,*args, **kwargs):
        objects=self.objects
        if "search_for" in kwargs:
            search_for=kwargs["search_for"]
            codeee=str(filter_number(search_for))
            objects=objects.filter(Q(name__contains=search_for) | Q(code=search_for) | Q(code=codeee) )
        if "account_code" in kwargs:
            account_code=kwargs["account_code"]
            objects=objects.filter(Q(bedehkar__code=account_code) | Q(bestankar__code=account_code)  )
        if "account_id" in kwargs:
            account_id=kwargs["account_id"]
            objects=objects.filter(Q(bedehkar_id=account_id) | Q(bestankar_id=account_id)  )
        if "parent_id" in kwargs:
            parent_id=kwargs["parent_id"]
            objects=objects.filter(parent_id=parent_id)  
        return objects.order_by('-event_datetime')
       
    def roots(self,*args, **kwargs):
        objects=self.objects.filter(parent_id=None)
        return objects.all()

    def cheque(self,*args, **kwargs):
        if "cheque_id" in kwargs and kwargs["cheque_id"] is not None:
            return self.objects.filter(pk=kwargs['cheque_id']).first()  
        if "pk" in kwargs and kwargs["pk"] is not None:
            return self.objects.filter(pk=kwargs['pk']).first() 
        if "id" in kwargs and kwargs["id"] is not None:
            return self.objects.filter(pk=kwargs['id']).first() 
        if "event_id" in kwargs and kwargs["event_id"] is not None:
            return self.objects.filter(pk=kwargs['event_id']).first() 
         
    def edit_cheque(self,*args, **kwargs):
        result,message,cheque=FAILED,"",None
        if not self.request.user.has_perm(APP_NAME+".add_financialevent"):
            message="دسترسی غیر مجاز"
            return result,message,cheque
    
        if "cheque_id" in kwargs and kwargs["cheque_id"] is not None:
            cheque=Cheque.objects.filter(pk=kwargs['cheque_id']).first()  
            if cheque is None:
                message="رویداد پیدا نشد."
                return result,message,cheque

        if "title" in kwargs and kwargs["title"] is not None:
            cheque.title=kwargs['title']


        if "status" in kwargs  :
            cheque.status=kwargs['status']


        if "amount" in kwargs and kwargs["amount"] is not None:
            cheque.amount=kwargs['amount']

        if "payment_method" in kwargs and kwargs["payment_method"] is not None:
            cheque.payment_method=kwargs['payment_method']

            
        if "shipping_fee" in kwargs and kwargs["shipping_fee"] is not None:
            cheque.shipping_fee=kwargs['shipping_fee'] 

        if "discount" in kwargs and kwargs["discount"] is not None:
            cheque.discount=kwargs['discount']
        if "tax_percentage" in kwargs and kwargs["tax_percentage"] is not None:
            cheque.tax_percentage=kwargs['tax_percentage']

        if "shipping_fee" in kwargs and kwargs["shipping_fee"] is not None:
            cheque.shipping_fee=kwargs['shipping_fee']

        if "bedehkar_id" in kwargs and kwargs["bedehkar_id"] is not None:
            cheque.bedehkar_id=kwargs['bedehkar_id']

        if "bestankar_id" in kwargs and kwargs["bestankar_id"] is not None:
            cheque.bestankar_id=kwargs['bestankar_id']
        
        if 'event_datetime1' in kwargs and kwargs['event_datetime'] is not None:
            event_datetime=kwargs["event_datetime"]
            cheque.event_datetime=event_datetime

           
            year=event_datetime[:2]
            if year=="13" or year=="14":
                event_datetime=PersianCalendar().to_gregorian(event_datetime)
            cheque.event_datetime=event_datetime

        # if 'start_datetime' in kwargs:
        #     start_datetime=kwargs["start_datetime"]
        #     cheque.start_datetime=start_datetime

           
        #     year=start_datetime[:2]
        #     if year=="13" or year=="14":
        #         start_datetime=PersianCalendar().to_gregorian(start_datetime)
        #     cheque.start_datetime=start_datetime

        # if 'end_datetime' in kwargs:
        #     end_datetime=kwargs["end_datetime"]
        #     cheque.end_datetime=end_datetime

           
        #     year=end_datetime[:2]
        #     if year=="13" or year=="14":
        #         end_datetime=PersianCalendar().to_gregorian(end_datetime)
        #     cheque.end_datetime=end_datetime

        result,message,cheque=cheque.save()

        return result,message,cheque
       
    def delete_all(self,*args, **kwargs):
        result,message=FAILED,""
        if not self.request.user.has_perm(APP_NAME+".delete_financialevent"):
            message="دسترسی غیر مجاز"
            return result,message
        cheques=Cheque.objects.all()
        cheques.delete()
        result=SUCCEED
        message='همه رویداد ها حذف شدند.'
        return result,message
    
    def add_cheque(self,*args,**kwargs):
        result,message,cheque=FAILED,"",None
        
        if len(Cheque.objects.filter(title=kwargs['title']))>0:
            message='عنوان تکراری' 
            return result,message,None
        
        if not self.request.user.has_perm(APP_NAME+".add_financialevent"):
            message="دسترسی غیر مجاز"
            return result,message,cheque

        cheque=Cheque()
        if 'bedehkar_id' in kwargs:
            cheque.bedehkar_id=kwargs["bedehkar_id"]
        if 'bestankar_id' in kwargs:
            cheque.bestankar_id=kwargs["bestankar_id"]

            
        if "shipping_fee" in kwargs and kwargs["shipping_fee"] is not None:
            cheque.shipping_fee=kwargs['shipping_fee'] 

        if "payment_method" in kwargs and kwargs["payment_method"] is not None:
            cheque.payment_method=kwargs['payment_method'] 

        if 'title' in kwargs and kwargs["title"] is not None:
            cheque.title=kwargs["title"]
        if 'status' in kwargs and kwargs["status"] is not None:
            cheque.status=kwargs["status"]
        if 'description' in kwargs and kwargs["description"] is not None:
            cheque.description=kwargs["description"]
        if 'parent_id' in kwargs and kwargs["parent_id"] is not None:
            if kwargs["parent_id"]>0:
                cheque.parent_id=kwargs["parent_id"]
        if 'color' in kwargs and kwargs["color"] is not None:
            cheque.color=kwargs["color"]
        if 'amount' in kwargs and kwargs["amount"] is not None:
            cheque.amount=kwargs["amount"]
        if 'priority' in kwargs and kwargs["priority"] is not None:
            cheque.priority=kwargs["priority"]
        if 'type' in kwargs and kwargs["type"] is not None:
            cheque.type=kwargs["type"]
        if 'event_datetime' in kwargs and kwargs["event_datetime"] is not None:
            event_datetime=kwargs["event_datetime"]
            cheque.event_datetime=event_datetime

           
            year=event_datetime[:2]
            if year=="13" or year=="14":
                event_datetime=PersianCalendar().to_gregorian(event_datetime)
            cheque.event_datetime=event_datetime

        # if 'start_datetime' in kwargs:
        #     start_datetime=kwargs["start_datetime"]
        #     cheque.start_datetime=start_datetime

           
        #     year=start_datetime[:2]
        #     if year=="13" or year=="14":
        #         start_datetime=PersianCalendar().to_gregorian(start_datetime)
        #     cheque.start_datetime=start_datetime

        # if 'end_datetime' in kwargs:
        #     end_datetime=kwargs["end_datetime"]
        #     cheque.end_datetime=end_datetime

           
        #     year=end_datetime[:2]
        #     if year=="13" or year=="14":
        #         end_datetime=PersianCalendar().to_gregorian(end_datetime)
        #     cheque.end_datetime=end_datetime

        (result,message,cheque)=cheque.save()
        return result,message,cheque
 